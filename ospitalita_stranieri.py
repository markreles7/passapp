from __future__ import annotations

import datetime as dt
import difflib
import glob
import json
import os
import re
import shutil
import threading
import unicodedata
import queue
import logging
import subprocess
import tempfile
from pathlib import Path
import tkinter as tk
from tkinter import messagebox, ttk

from app_config import load_config, resolve_path

try:
    import xlrd  # type: ignore
except ImportError:  # pragma: no cover
    xlrd = None

try:
    import openpyxl  # type: ignore
except ImportError:  # pragma: no cover
    openpyxl = None

APP_CONFIG = load_config()
PATHS = APP_CONFIG["paths"]
UI_CONFIG = APP_CONFIG["ui"]
OSPITALITA_UI = UI_CONFIG["modules"]["ospitalita"]

FOLDER_OSPITALITA = PATHS["ospitalita_network_folder"]
FILE_PATTERNS = list(PATHS["ospitalita_patterns"])
WORK_COPY_DIR = resolve_path("data/workcopies/ospitalita")

BG = "#EEF4EF"
BG2 = "#E1ECE4"
SURFACE = "#FFFFFF"
BORDER = "#C9D8CC"
ACCENT = "#1F6F4A"
ACCENT_DARK = "#174F35"
TEXT = "#132018"
TEXT_MUTED = "#4E6354"
TEXT_DIM = "#7C8F81"
WARNING = "#AD6A0F"
DANGER = "#B43A30"

LOG_FILE = resolve_path(PATHS["log_file"])
try:
    LOG_FILE.parent.mkdir(parents=True, exist_ok=True)
except OSError:
    pass

logger = logging.getLogger(__name__)
if not logger.handlers:
    try:
        _handler = logging.FileHandler(LOG_FILE, encoding="utf-8")
        _handler.setFormatter(logging.Formatter("%(asctime)s | %(levelname)s | %(name)s | %(message)s"))
        logger.addHandler(_handler)
    except OSError:
        pass
logger.setLevel(logging.INFO)
logger.propagate = False

FIELD_ALIASES = {
    "protocollo": ["protocollo", "n protocollo", "num protocollo", "prot"],
    "data_presentazione": [
        "data presentazione",
        "data comunicazione",
        "data denuncia",
        "data",
        "presentazione",
    ],
    "denunciante_dichiarante": [
        "denunciante",
        "dichiarante",
        "comunicante",
        "richiedente",
        "segnalante",
    ],
    "cittadino_ospitato": [
        "ospitato",
        "cittadino ospitato",
        "straniero ospitato",
        "ospite",
        "generalita ospitato",
        "cognome",
        "nome",
        "nominativo",
    ],
    "indirizzo": ["indirizzo", "via", "residenza", "domicilio", "alloggio", "luogo ospitalita"],
    "tipo_comunicazione": ["tipo comunicazione", "tipologia", "tipo", "comunicazione", "motivo"],
}


def _norm(value) -> str:
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return text.strip().lower()


def _text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, dt.date):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _parse_sort_date(value):
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y"):
        try:
            return dt.datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def _protocol_sort_key(value):
    text = _text(value)
    if not text:
        return (1, tuple(), "")
    nums = re.findall(r"\d+", text)
    if nums:
        return (0, tuple(int(n) for n in nums), _norm(text))
    return (1, tuple(), _norm(text))


def _extract_year_from_name(name: str) -> int:
    years = [int(m.group(0)) for m in re.finditer(r"(?:19|20)\d{2}", name)]
    return max(years) if years else -1


def _file_sort_key(path: str):
    base = os.path.basename(path)
    year = _extract_year_from_name(base)
    try:
        mtime = os.path.getmtime(path)
    except OSError:
        mtime = 0.0
    return (year, mtime, base.lower())


def _looks_empty_row(row) -> bool:
    return all(_norm(cell) == "" for cell in row)


def _tokens(text: str) -> set[str]:
    return {t for t in re.split(r"[^a-z0-9]+", text) if t}


def _match_score(header_text: str, alias: str) -> float:
    h = _norm(header_text)
    a = _norm(alias)
    if not h or not a:
        return 0.0
    if h == a:
        return 1.0
    if a in h:
        return 0.92
    if h in a:
        return 0.76

    ht = _tokens(h)
    at = _tokens(a)
    overlap = 0.0
    if ht and at:
        overlap = len(ht & at) / max(len(at), 1)

    ratio = difflib.SequenceMatcher(None, h, a).ratio()
    return max(overlap * 0.9, ratio * 0.8)


def _best_alias_score(header_text: str, aliases: list[str]) -> float:
    return max((_match_score(header_text, alias) for alias in aliases), default=0.0)


def _header_score(row) -> float:
    headers = [_norm(c) for c in row if _norm(c)]
    if not headers:
        return 0.0
    score = 0.0
    for head in headers:
        score += max((_best_alias_score(head, aliases) for aliases in FIELD_ALIASES.values()), default=0.0)
    return score


def _find_header_row(rows) -> int:
    best_idx = 0
    best_score = -1.0
    limit = min(30, len(rows))
    for idx in range(limit):
        row = rows[idx]
        score = _header_score(row)
        if score > best_score:
            best_idx = idx
            best_score = score
    return best_idx


def _map_columns(headers) -> dict[str, int | None]:
    mapping: dict[str, int | None] = {}
    used: set[int] = set()
    for field, aliases in FIELD_ALIASES.items():
        best_idx = None
        best_score = 0.0
        for idx, head in enumerate(headers):
            if idx in used:
                continue
            score = _best_alias_score(head, aliases)
            if score > best_score:
                best_score = score
                best_idx = idx
        if best_idx is not None and best_score >= 0.45:
            mapping[field] = best_idx
            used.add(best_idx)
        else:
            mapping[field] = None
    return mapping


def _load_rows_xls(path: str):
    if xlrd is None:
        logger.error("Import .xls non possibile: libreria xlrd non disponibile (%s)", path)
        raise RuntimeError("Libreria 'xlrd' non disponibile per leggere file .xls")
    book = xlrd.open_workbook(path)

    best_sheet_idx = 0
    best_sheet_score = -1.0
    best_header_idx = 0
    for i in range(book.nsheets):
        sh = book.sheet_by_index(i)
        sample = []
        limit = min(sh.nrows, 35)
        for r in range(limit):
            sample.append([sh.cell_value(r, c) for c in range(sh.ncols)])
        if not sample:
            continue
        hidx = _find_header_row(sample)
        score = _header_score(sample[hidx]) if hidx < len(sample) else 0.0
        if score > best_sheet_score:
            best_sheet_score = score
            best_sheet_idx = i
            best_header_idx = hidx

    sheet = book.sheet_by_index(best_sheet_idx)
    rows = []
    for r in range(sheet.nrows):
        parsed_row = []
        for c in range(sheet.ncols):
            cell = sheet.cell(r, c)
            value = cell.value
            if cell.ctype == getattr(xlrd, "XL_CELL_DATE", 3):
                try:
                    value = xlrd.xldate_as_datetime(value, book.datemode)
                except Exception:
                    pass
            parsed_row.append(value)
        rows.append(parsed_row)
    meta = {
        "active_sheet": sheet.name,
        "active_sheet_index": best_sheet_idx,
        "header_hint": best_header_idx + 1,
        "mode": "best_sheet",
    }
    return rows, meta


def _load_rows_xlsx(path: str):
    if openpyxl is None:
        logger.error("Import .xlsx non possibile: libreria openpyxl non disponibile (%s)", path)
        raise RuntimeError("Libreria 'openpyxl' non disponibile per leggere file .xlsx")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    active_ws = wb.active
    active_title = active_ws.title

    sheet_rows: list[tuple[str, list[list]]]=[]
    for ws in wb.worksheets:
        ws_rows = [list(r) for r in ws.iter_rows(values_only=True)]
        sheet_rows.append((ws.title, ws_rows))

    selected_title = active_title
    selected_rows = []
    best_score = -1.0
    for title, rows in sheet_rows:
        if not rows:
            continue
        hidx = _find_header_row(rows)
        score = _header_score(rows[hidx]) if hidx < len(rows) else 0.0
        if title == active_title:
            selected_rows = rows
            best_score = score
        if score > best_score + 0.25:
            selected_title = title
            selected_rows = rows
            best_score = score

    if not selected_rows:
        selected_rows = [list(r) for r in active_ws.iter_rows(values_only=True)]

    rows = selected_rows
    wb.close()
    meta = {
        "active_sheet": selected_title,
        "workbook_active_sheet": active_title,
        "mode": "active_or_best",
    }
    return rows, meta


def _load_rows(path: str):
    lower = path.lower()
    if lower.endswith(".xlsx"):
        return _load_rows_xlsx(path)
    if lower.endswith(".xls"):
        return _load_rows_xls(path)
    raise RuntimeError("Formato file non supportato")


def _extract_records(path: str):
    rows, meta = _load_rows(path)
    if not rows:
        return [], "Nessuna riga trovata"

    header_idx = _find_header_row(rows)
    headers = rows[header_idx]
    mapping = _map_columns(headers)
    records = []

    def pick(row, key):
        idx = mapping.get(key)
        if idx is None or idx >= len(row):
            return ""
        return _text(row[idx])

    def progressivo(row):
        if not row:
            return ""
        value = _text(row[0] if len(row) > 0 else "")
        marker = _norm(value)
        if marker in {"", "n", "n.", "nr", "numero", "num", "progressivo", "n°"}:
            return ""
        return value

    def is_continuation_record(record: dict) -> bool:
        return not (record["progressivo"] or record["protocollo"] or record["data_presentazione"])

    def merge_value(current: str, extra: str) -> str:
        cur = _text(current).strip()
        add = _text(extra).strip()
        if not add or _norm(add) in {"-", "—"}:
            return cur
        if not cur or _norm(cur) in {"-", "—"}:
            return add
        parts = [p.strip() for p in cur.split(" | ") if p.strip()]
        if add not in parts:
            parts.append(add)
        return " | ".join(parts)

    for row in rows[header_idx + 1 :]:
        if _looks_empty_row(row):
            continue

        record = {
            "progressivo": progressivo(row),
            "protocollo": pick(row, "protocollo"),
            "data_presentazione": pick(row, "data_presentazione"),
            "denunciante_dichiarante": pick(row, "denunciante_dichiarante"),
            "cittadino_ospitato": pick(row, "cittadino_ospitato"),
            "indirizzo": pick(row, "indirizzo"),
            "tipo_comunicazione": pick(row, "tipo_comunicazione"),
            "source": os.path.basename(path),
            "sheet": meta.get("active_sheet", ""),
        }

        has_any_data = (
            record["progressivo"]
            or record["protocollo"]
            or record["data_presentazione"]
            or record["denunciante_dichiarante"]
            or record["cittadino_ospitato"]
            or record["indirizzo"]
            or record["tipo_comunicazione"]
        )
        if not has_any_data:
            continue
        if is_continuation_record(record) and records:
            prev = records[-1]
            for key in ("denunciante_dichiarante", "cittadino_ospitato", "indirizzo", "tipo_comunicazione"):
                prev[key] = merge_value(prev.get(key, ""), record.get(key, ""))
            continue
        records.append(record)

    mapped_cols = ", ".join(f"{k}:{v}" for k, v in mapping.items())
    active_info = meta.get("active_sheet", "")
    workbook_active = meta.get("workbook_active_sheet")
    if workbook_active and workbook_active != active_info:
        sheet_info = f"sheet={active_info} (active workbook={workbook_active})"
    else:
        sheet_info = f"sheet={active_info}"
    insight = f"{os.path.basename(path)} -> {sheet_info} header row {header_idx + 1} ({mapped_cols})"
    return records, insight


def _list_input_files() -> list[str]:
    files: list[str] = []
    for pattern in FILE_PATTERNS:
        full = os.path.join(FOLDER_OSPITALITA, pattern)
        files.extend(glob.glob(full))
    dedup = sorted(set(files), key=_file_sort_key, reverse=True)
    return dedup


def _detect_layout(path: str) -> dict:
    rows, meta = _load_rows(path)
    if not rows:
        return {"sheet": meta.get("active_sheet", ""), "mapping": {}, "header_idx": 0}
    header_idx = _find_header_row(rows)
    headers = rows[header_idx] if header_idx < len(rows) else []
    mapping = _map_columns(headers)
    return {"sheet": meta.get("active_sheet", ""), "mapping": mapping, "header_idx": header_idx}


class OspitalitaStranieriFrame(tk.Frame):
    def __init__(self, parent, controller):
        super().__init__(parent, bg=BG)
        self.controller = controller

        self.all_records: list[dict] = []
        self.filtered: list[dict] = []
        self.sort_col: str | None = None
        self.sort_asc = True
        self.last_insights: list[str] = []
        self._load_seq = 0
        self._all_records_loaded: list[dict] = []
        self._files_loaded: list[str] = []
        self._source_files: list[str] = []
        self._primary_source_file: str | None = None
        self._working_copy_file: Path | None = None
        self._pending_new_records: list[dict] = []
        self.year_var = tk.StringVar(value="Tutti")
        self._last_year_value = "Tutti"

        self._ttk_style = ttk.Style(self)
        self._setup_styles()
        self._build_ui()
        self.after(200, self.carica_dati)

    def on_show(self):
        self.controller.title(OSPITALITA_UI["title"])

    def _setup_styles(self):
        try:
            self._ttk_style.theme_use("clam")
        except tk.TclError:
            pass

        self._ttk_style.configure(
            "Osp.Treeview",
            background=SURFACE,
            fieldbackground=SURFACE,
            foreground=TEXT,
            rowheight=34,
            font=("Segoe UI", 10),
            bordercolor=BORDER,
            relief="flat",
        )
        self._ttk_style.configure(
            "Osp.Treeview.Heading",
            background="#E7F0EA",
            foreground="#355343",
            font=("Segoe UI", 10, "bold"),
            relief="flat",
            bordercolor=BORDER,
        )
        self._ttk_style.map("Osp.Treeview", background=[("selected", "#CEE5D6")], foreground=[("selected", "#173E2B")])
        self._ttk_style.map("Osp.Treeview.Heading", background=[("active", BORDER)])

    def _build_ui(self):
        hdr = tk.Frame(self, bg=SURFACE)
        hdr.pack(fill="x", side="top")
        tk.Frame(hdr, bg=BORDER, height=1).pack(fill="x", side="bottom")

        inner = tk.Frame(hdr, bg=SURFACE)
        inner.pack(fill="x", padx=24, pady=14)

        left = tk.Frame(inner, bg=SURFACE)
        left.pack(side="left")

        badge = tk.Label(left, text="🏠", bg=ACCENT, fg="white", font=("Segoe UI", 16), width=2, padx=6, pady=4)
        badge.pack(side="left", padx=(0, 12))

        title_box = tk.Frame(left, bg=SURFACE)
        title_box.pack(side="left")
        tk.Label(
            title_box,
            text=OSPITALITA_UI["menu_title"],
            bg=SURFACE,
            fg=TEXT,
            font=("Segoe UI", 15, "bold"),
        ).pack(anchor="w")
        tk.Label(
            title_box,
            text="Polizia Locale",
            bg=SURFACE,
            fg=ACCENT,
            font=("Segoe UI", 10, "bold"),
        ).pack(anchor="w", pady=(2, 0))

        actions = tk.Frame(inner, bg=SURFACE)
        actions.pack(side="right")

        self.lbl_tot = self._stat(actions, "📄 Totale: —")
        self.lbl_files = self._stat(actions, "📂 File: —")

        tk.Button(
            actions,
            text="← Torna al Menu Principale",
            bg=BG2,
            fg=TEXT_MUTED,
            font=("Segoe UI", 10, "bold"),
            relief="flat",
            cursor="hand2",
            activebackground=BORDER,
            highlightbackground=BORDER,
            highlightthickness=1,
            padx=12,
            pady=8,
            command=self.torna_menu,
        ).pack(side="right", padx=(8, 0))

        body = tk.Frame(self, bg=BG)
        body.pack(fill="both", expand=True, padx=24, pady=(14, 18))

        search_row = tk.Frame(body, bg=BG)
        search_row.pack(fill="x")

        search_box = tk.Frame(search_row, bg=SURFACE, highlightbackground=BORDER, highlightthickness=1)
        search_box.pack(side="left", fill="x", expand=True)

        tk.Label(search_box, text="🔍", bg=SURFACE, font=("Segoe UI", 13), padx=10).pack(side="left")
        self.search_var = tk.StringVar()
        self.search_var.trace_add("write", lambda *_: self.applica_filtro())
        self.search_entry = tk.Entry(
            search_box,
            textvariable=self.search_var,
            font=("Segoe UI", 12),
            relief="flat",
            bg=SURFACE,
            fg=TEXT,
            insertbackground=ACCENT,
            bd=0,
        )
        self.search_entry.pack(side="left", fill="x", expand=True, ipady=8)

        tk.Button(
            search_box,
            text="✕",
            bg=SURFACE,
            fg=TEXT_MUTED,
            font=("Segoe UI", 10),
            relief="flat",
            cursor="hand2",
            activebackground=BG2,
            command=lambda: self.search_var.set(""),
        ).pack(side="right", padx=8)

        self.btn_refresh = tk.Button(
            search_row,
            text="↻  Aggiorna",
            bg=ACCENT,
            fg="white",
            font=("Segoe UI", 10, "bold"),
            relief="flat",
            cursor="hand2",
            activebackground=ACCENT_DARK,
            padx=16,
            pady=9,
            command=self.carica_dati,
        )
        self.btn_refresh.pack(side="left", padx=(10, 0))

        self.btn_new = tk.Button(
            search_row,
            text="+ Nuovo nominativo",
            bg=ACCENT,
            fg="white",
            font=("Segoe UI", 10, "bold"),
            relief="flat",
            cursor="hand2",
            activebackground=ACCENT_DARK,
            padx=14,
            pady=9,
            command=self.nuovo_nominativo,
            state="disabled",
        )
        self.btn_new.pack(side="left", padx=(8, 0))

        self.btn_edit = tk.Button(
            search_row,
            text="Modifica selezionato",
            bg=ACCENT,
            fg="white",
            font=("Segoe UI", 10, "bold"),
            relief="flat",
            cursor="hand2",
            activebackground=ACCENT_DARK,
            padx=14,
            pady=9,
            command=self.modifica_selezionato,
            state="disabled",
        )
        self.btn_edit.pack(side="left", padx=(8, 0))

        self.btn_save_changes = tk.Button(
            search_row,
            text="SALVA MODIFICHE",
            bg=WARNING,
            fg="white",
            font=("Segoe UI", 10, "bold"),
            relief="flat",
            cursor="hand2",
            activebackground="#915B10",
            padx=14,
            pady=9,
            command=self.salva_modifiche,
            state="disabled",
        )
        self.btn_save_changes.pack(side="left", padx=(8, 0))

        year_wrap = tk.Frame(search_row, bg=BG)
        year_wrap.pack(side="left", padx=(10, 0))
        tk.Label(
            year_wrap,
            text="Anno",
            bg=BG,
            fg=TEXT_MUTED,
            font=("Segoe UI", 9, "bold"),
        ).pack(side="left", padx=(0, 6))
        self.year_combo = ttk.Combobox(
            year_wrap,
            textvariable=self.year_var,
            values=("Tutti",),
            state="readonly",
            width=10,
            font=("Segoe UI", 10),
        )
        self.year_combo.pack(side="left")
        self.year_combo.bind("<<ComboboxSelected>>", self._on_year_change)

        tk.Label(
            body,
            text="Ricerca per n. progressivo, protocollo, nominativo, indirizzo o data.",
            bg=BG,
            fg=TEXT_MUTED,
            font=("Segoe UI", 9),
            anchor="w",
        ).pack(fill="x", pady=(8, 0))

        self.lbl_status = tk.Label(body, text="Pronto.", bg=BG, fg=TEXT_DIM, font=("Segoe UI", 8), anchor="w")
        self.lbl_status.pack(fill="x", pady=(6, 8))

        table_shell = tk.Frame(body, bg=BG2, highlightbackground=BORDER, highlightthickness=1)
        table_shell.pack(fill="both", expand=True)

        table_wrap = tk.Frame(table_shell, bg=SURFACE)
        table_wrap.pack(fill="both", expand=True, padx=1, pady=1)

        vsb = ttk.Scrollbar(table_wrap, orient="vertical")
        vsb.pack(side="right", fill="y")

        self.columns = (
            "progressivo",
            "protocollo",
            "denunciante_dichiarante",
            "cittadino_ospitato",
            "data_presentazione",
            "dettagli",
            "source",
        )
        self.tree = ttk.Treeview(
            table_wrap,
            columns=self.columns,
            show="headings",
            yscrollcommand=vsb.set,
            selectmode="browse",
            style="Osp.Treeview",
        )
        vsb.config(command=self.tree.yview)

        headers = {
            "progressivo": ("N°", 56, "center"),
            "protocollo": ("Protocollo", 88, "center"),
            "denunciante_dichiarante": ("Denunciante", 330, "w"),
            "cittadino_ospitato": ("Straniero", 320, "w"),
            "data_presentazione": ("Data", 95, "center"),
            "dettagli": ("Dettaglio Ospitalita", 290, "w"),
            "source": ("File", 200, "w"),
        }
        for col, (label, width, anchor) in headers.items():
            self.tree.heading(col, text=label, command=lambda c=col: self.ordina(c))
            self.tree.column(col, width=width, anchor=anchor, minwidth=40)

        self.tree.pack(fill="both", expand=True)
        self.tree.bind("<Double-1>", self.modifica_selezionato)

    def _stat(self, parent, text):
        wrap = tk.Frame(parent, bg=BG2, highlightbackground=BORDER, highlightthickness=1)
        wrap.pack(side="right", padx=4)
        lbl = tk.Label(wrap, text=text, bg=BG2, fg=TEXT_MUTED, font=("Segoe UI", 9, "bold"), padx=10, pady=5)
        lbl.pack()
        return lbl

    @staticmethod
    def _year_from_label(text: str) -> str:
        year = _extract_year_from_name(text)
        return str(year) if year > 0 else ""

    def _file_year(self, path: str) -> str:
        return self._year_from_label(os.path.basename(path))

    def _record_year(self, rec: dict) -> str:
        return self._year_from_label(str(rec.get("source", "")))

    def _refresh_year_choices(self):
        years = sorted(
            {self._file_year(path) for path in self._files_loaded if self._file_year(path)},
            reverse=True,
        )
        values = ["Tutti"] + years if years else ["Tutti"]
        self.year_combo.configure(values=values)

        current = self.year_var.get().strip()
        if current in values and current != "Tutti":
            selected = current
        else:
            selected = years[0] if years else "Tutti"

        self.year_var.set(selected)
        self._last_year_value = selected

    def _apply_year_context(self):
        selected_year = self.year_var.get().strip()
        selected_files = list(self._files_loaded)
        selected_records = list(self._all_records_loaded)

        if selected_year and selected_year != "Tutti":
            selected_files = [path for path in self._files_loaded if self._file_year(path) == selected_year]
            source_names = {os.path.basename(path) for path in selected_files}
            selected_records = [rec for rec in self._all_records_loaded if str(rec.get("source", "")) in source_names]

        self._source_files = selected_files
        self.all_records = selected_records
        if selected_year == "Tutti" and len(selected_files) != 1:
            self._primary_source_file = None
        else:
            self._primary_source_file = selected_files[0] if selected_files else None

        self._prepare_working_copy()
        self.btn_new.config(state="normal" if self._working_copy_file else "disabled")
        self.btn_edit.config(state="normal" if self._working_copy_file else "disabled")
        self.lbl_tot.config(text=f"Totale: {len(self.all_records)}")
        self.lbl_files.config(text=f"File: {len(self._source_files)}")

    def _on_year_change(self, _event=None):
        next_year = self.year_var.get().strip() or "Tutti"
        prev_year = self._last_year_value
        if next_year == prev_year:
            return

        if self._pending_new_records and not self._confirm_pending_before_context_switch():
            self.year_var.set(prev_year)
            return

        self._last_year_value = next_year
        self._pending_new_records = []
        self.btn_save_changes.config(state="disabled")
        self._apply_year_context()
        self.applica_filtro()

    def carica_dati(self, force: bool = False):
        if not force and self._pending_new_records:
            if not self._confirm_pending_before_context_switch():
                return

        self.btn_refresh.config(state="disabled", text="Caricamento…")
        self.lbl_status.config(text=f"Lettura da {FOLDER_OSPITALITA} ...")
        self.tree.delete(*self.tree.get_children())
        self._load_seq += 1
        current_seq = self._load_seq
        result_q: queue.Queue = queue.Queue()

        def worker():
            files = []
            errors = []
            all_records: list[dict] = []
            insights = []

            if not os.path.isdir(FOLDER_OSPITALITA):
                errors.append("Percorso di rete non raggiungibile.")
                logger.warning("Percorso ospitalita non raggiungibile: %s", FOLDER_OSPITALITA)
            else:
                files = _list_input_files()
                for path in files:
                    try:
                        recs, insight = _extract_records(path)
                        all_records.extend(recs)
                        insights.append(insight)
                    except Exception as exc:
                        logger.exception("Errore importazione file ospitalita: %s", path)
                        errors.append(f"{os.path.basename(path)}: {exc}")

            result_q.put((all_records, files, errors, insights))

        def poll():
            if current_seq != self._load_seq:
                return
            if not self.winfo_exists():
                return
            try:
                all_records, files, errors, insights = result_q.get_nowait()
            except queue.Empty:
                try:
                    self.after(60, poll)
                except tk.TclError:
                    pass
                return

            if current_seq != self._load_seq or not self.winfo_exists():
                return
            self._post_carica(all_records, files, errors, insights)

        threading.Thread(target=worker, daemon=True).start()
        self.after(60, poll)

    def _post_carica(self, records, files, errors, insights):
        self.btn_refresh.config(state="normal", text="↻  Aggiorna")
        self._all_records_loaded = list(records)
        self.all_records = list(records)
        self.last_insights = insights

        if not files:
            self.lbl_status.config(text="Nessun file trovato o cartella non raggiungibile.")
            if errors:
                messagebox.showwarning(
                    "Origine dati non disponibile",
                    f"Impossibile leggere i file da:\n{FOLDER_OSPITALITA}\n\nDettagli:\n" + "\n".join(errors[:8]),
                )
            self.lbl_tot.config(text="Totale: 0")
            self.lbl_files.config(text="File: 0")
            self.filtered = []
            self._all_records_loaded = []
            self._files_loaded = []
            self._source_files = []
            self._primary_source_file = None
            self._working_copy_file = None
            self._pending_new_records = []
            self.btn_new.config(state="disabled")
            self.btn_edit.config(state="disabled")
            self.btn_save_changes.config(state="disabled")
            self.year_combo.configure(values=("Tutti",))
            self.year_var.set("Tutti")
            self._last_year_value = "Tutti"
            return

        if errors:
            logger.error("Errori importazione ospitalita: %s", " | ".join(errors))
            messagebox.showwarning(
                "Alcuni file non letti",
                "Sono presenti errori di importazione:\n\n" + "\n".join(errors[:10]),
            )

        self._files_loaded = list(files)
        self._pending_new_records = []
        self._refresh_year_choices()
        self._apply_year_context()

        self.lbl_tot.config(text=f"Totale: {len(self.all_records)}")
        self.lbl_files.config(text=f"File: {len(self._source_files)}")

        insight_text = insights[0] if insights else "Nessun mapping header disponibile."
        selected_year = self.year_var.get().strip()
        year_label = f"Anno {selected_year}" if selected_year and selected_year != "Tutti" else "Tutti gli anni"
        self.lbl_status.config(
            text=(
                f"{year_label} · Caricati {len(self._source_files)} file · "
                f"{len(self.all_records)} record · {dt.datetime.now().strftime('%H:%M:%S')} | {insight_text}"
            )
        )
        self.btn_new.config(state="normal" if self._working_copy_file else "disabled")
        self.btn_edit.config(state="normal" if self._working_copy_file else "disabled")
        self.btn_save_changes.config(state="disabled")
        self.applica_filtro()

    def applica_filtro(self):
        query = self.search_var.get().strip().lower()
        if not query:
            self.filtered = list(self.all_records)
        else:
            out = []
            for rec in self.all_records:
                searchable = " ".join(
                    [
                        rec.get("progressivo", ""),
                        rec.get("protocollo", ""),
                        rec.get("data_presentazione", ""),
                        rec.get("denunciante_dichiarante", ""),
                        rec.get("cittadino_ospitato", ""),
                        rec.get("indirizzo", ""),
                        rec.get("tipo_comunicazione", ""),
                    ]
                ).lower()
                if query in searchable:
                    out.append(rec)
            self.filtered = out
        self._popola_tabella(self.filtered)

    @staticmethod
    def _compact_values(*values) -> str:
        parts: list[str] = []
        for value in values:
            text = _text(value).strip()
            if not text or _norm(text) in {"-", "—"}:
                continue
            if text not in parts:
                parts.append(text)
        return " | ".join(parts) if parts else "-"

    def _popola_tabella(self, records):
        self.tree.delete(*self.tree.get_children())
        for idx, rec in enumerate(records):
            tag = "odd" if idx % 2 else "even"
            denunciante = self._compact_values(rec.get("denunciante_dichiarante", ""))
            straniero = self._compact_values(rec.get("cittadino_ospitato", ""))
            dettagli = self._compact_values(rec.get("tipo_comunicazione", ""), rec.get("indirizzo", ""))
            self.tree.insert(
                "",
                "end",
                iid=f"rec-{idx}",
                tags=(tag,),
                values=(
                    rec.get("progressivo", "-") or "-",
                    rec.get("protocollo", "-") or "-",
                    denunciante,
                    straniero,
                    rec.get("data_presentazione", "-") or "-",
                    dettagli,
                    rec.get("source", "-") or "-",
                ),
            )

        self.tree.tag_configure("odd", background="#F5FAF6")
        self.tree.tag_configure("even", background=SURFACE)

    def ordina(self, col):
        if self.sort_col == col:
            self.sort_asc = not self.sort_asc
        else:
            self.sort_col = col
            self.sort_asc = True

        self.filtered.sort(
            key=lambda r: self._sort_key(r, col),
            reverse=not self.sort_asc,
        )
        self._popola_tabella(self.filtered)

        base = {
            "progressivo": "N°",
            "protocollo": "Protocollo",
            "denunciante_dichiarante": "Denunciante",
            "cittadino_ospitato": "Straniero",
            "data_presentazione": "Data",
            "dettagli": "Dettaglio Ospitalita",
            "source": "File",
        }
        for c in self.columns:
            arrow = (" ▲" if self.sort_asc else " ▼") if c == col else ""
            self.tree.heading(c, text=base[c] + arrow)

    def _sort_key(self, record, col):
        value = record.get(col, "")
        if col == "denunciante_dichiarante":
            value = self._compact_values(record.get("denunciante_dichiarante", ""))
        if col == "cittadino_ospitato":
            value = self._compact_values(record.get("cittadino_ospitato", ""))
        if col == "dettagli":
            value = self._compact_values(record.get("tipo_comunicazione", ""), record.get("indirizzo", ""))
        if col == "data_presentazione":
            parsed = _parse_sort_date(value)
            if parsed is not None:
                return (0, parsed, "")
            return (1, dt.date.min, _norm(value))
        if col == "progressivo":
            return _protocol_sort_key(value)
        if col == "protocollo":
            return _protocol_sort_key(value)
        return (0, _norm(value))

    def _prepare_working_copy(self):
        self._working_copy_file = None
        if not self._primary_source_file:
            return
        try:
            WORK_COPY_DIR.mkdir(parents=True, exist_ok=True)
            suffix = Path(self._primary_source_file).suffix or ".xlsx"
            fd, temp_name = tempfile.mkstemp(prefix="ospitalita_", suffix=suffix, dir=str(WORK_COPY_DIR))
            os.close(fd)
            temp_path = Path(temp_name)
            shutil.copy2(self._primary_source_file, temp_path)
            self._working_copy_file = temp_path
        except OSError:
            logger.exception("Impossibile creare copia di lavoro ospitalita")
            messagebox.showwarning(
                "Copia di lavoro non disponibile",
                "Non e stato possibile creare la copia di lavoro del registro.",
            )

    def _confirm_pending_before_context_switch(self) -> bool:
        answer = messagebox.askyesnocancel(
            "Modifiche non salvate",
            "Sono presenti modifiche non salvate.\nVuoi salvarle prima di continuare?",
            parent=self,
        )
        if answer is None:
            return False
        if answer:
            return self.salva_modifiche(trigger_reload=False)
        self._discard_pending_changes()
        return True

    def torna_menu(self):
        if self._pending_new_records:
            if not self._confirm_pending_before_context_switch():
                return
        self.controller.show_frame("MainMenuFrame")

    def _discard_pending_changes(self):
        for pending in reversed(self._pending_new_records):
            mode = pending.get("mode")
            progressivo = str(pending.get("progressivo", ""))
            source = str(pending.get("source", "")).strip()
            if mode in {"update_slot", "update_existing"}:
                target = self._find_record_by_progressivo(progressivo, source_name=source or None)
                original = pending.get("original_snapshot")
                if target is not None and isinstance(original, dict):
                    target.clear()
                    target.update(original)
            elif mode == "append":
                self.all_records = [
                    rec for rec in self.all_records
                    if not (
                        rec.get("_pending")
                        and str(rec.get("progressivo", "")).strip() == progressivo
                        and str(rec.get("source", "")).strip() == source
                    )
                ]

        self._pending_new_records = []
        self.btn_save_changes.config(state="disabled")
        self.applica_filtro()

    @staticmethod
    def _record_progressivo_int(record: dict) -> int | None:
        nums = re.findall(r"\d+", str(record.get("progressivo", "")))
        if not nums:
            return None
        try:
            return int(nums[0])
        except ValueError:
            return None

    @staticmethod
    def _blankish(value) -> bool:
        normalized = _norm(value)
        return normalized in {"", "-", "—", "none", "nan"}

    def _current_source_name(self) -> str:
        if not self._primary_source_file:
            return ""
        return os.path.basename(self._primary_source_file).strip()

    def _is_empty_slot(self, record: dict) -> bool:
        fields = (
            record.get("protocollo", ""),
            record.get("data_presentazione", ""),
            record.get("denunciante_dichiarante", ""),
            record.get("cittadino_ospitato", ""),
            record.get("tipo_comunicazione", ""),
            record.get("indirizzo", ""),
        )
        return all(self._blankish(v) for v in fields)

    def _find_record_by_progressivo(self, progressivo: str, source_name: str | None = None) -> dict | None:
        try:
            target = int(progressivo)
        except (TypeError, ValueError):
            return None

        source_ref = (source_name or self._current_source_name()).strip()
        for rec in self.all_records:
            if source_ref and str(rec.get("source", "")).strip() != source_ref:
                continue
            value = self._record_progressivo_int(rec)
            if value == target:
                return rec
        return None

    def _reserve_progressivo_slot(self) -> tuple[str, dict | None, dict | None, str]:
        source_ref = self._current_source_name()
        numeric_records: list[tuple[int, dict]] = []
        for rec in self.all_records:
            if source_ref and str(rec.get("source", "")).strip() != source_ref:
                continue
            prog = self._record_progressivo_int(rec)
            if prog is not None:
                numeric_records.append((prog, rec))

        for prog, rec in sorted(numeric_records, key=lambda item: item[0]):
            if rec.get("_pending"):
                continue
            if self._is_empty_slot(rec):
                return str(prog), rec, dict(rec), "update_slot"

        max_prog = max((prog for prog, _ in numeric_records), default=0)
        return str(max_prog + 1), None, None, "append"

    def _selected_record(self) -> dict | None:
        sel = self.tree.selection()
        if not sel:
            return None
        iid = sel[0]
        if not iid.startswith("rec-"):
            return None
        try:
            idx = int(iid.split("-", 1)[1])
        except ValueError:
            return None
        if idx < 0 or idx >= len(self.filtered):
            return None
        return self.filtered[idx]

    @staticmethod
    def _split_display_name_address(text: str) -> tuple[str, str]:
        raw = _text(text).strip()
        if not raw:
            return "", ""
        parts = [part.strip() for part in raw.split("|") if part.strip()]
        if len(parts) >= 2:
            return parts[0], parts[1]
        return raw, ""

    def _upsert_pending_record(self, pending: dict):
        source = str(pending.get("source", "")).strip()
        progressivo = str(pending.get("progressivo", "")).strip()
        for idx, existing in enumerate(self._pending_new_records):
            if (
                str(existing.get("source", "")).strip() == source
                and str(existing.get("progressivo", "")).strip() == progressivo
            ):
                if existing.get("original_snapshot") and not pending.get("original_snapshot"):
                    pending["original_snapshot"] = existing.get("original_snapshot")
                self._pending_new_records[idx] = pending
                return
        self._pending_new_records.append(pending)

    def modifica_selezionato(self, _event=None):
        if not self._working_copy_file:
            messagebox.showwarning(
                "Modifica non disponibile",
                "Per modificare i record seleziona un anno specifico.",
            )
            return
        record = self._selected_record()
        if record is None:
            messagebox.showinfo("Nessuna selezione", "Seleziona un nominativo da modificare.")
            return
        self.nuovo_nominativo(record_to_edit=record)

    def nuovo_nominativo(self, record_to_edit: dict | None = None):
        if not self._working_copy_file:
            messagebox.showwarning("Operazione non disponibile", "Nessuna copia di lavoro disponibile.")
            return

        is_edit = record_to_edit is not None
        win = tk.Toplevel(self)
        win.title("Modifica nominativo ospitalita" if is_edit else "Nuovo nominativo ospitalita")
        win.configure(bg=BG)
        win.resizable(False, False)
        win.grab_set()

        body = tk.Frame(win, bg=SURFACE, highlightbackground=BORDER, highlightthickness=1)
        body.pack(fill="both", expand=True, padx=14, pady=14)

        vars_map = {
            "protocollo": tk.StringVar(value=_text((record_to_edit or {}).get("protocollo", ""))),
            "denunciante_nome": tk.StringVar(),
            "denunciante_indirizzo": tk.StringVar(),
            "straniero_nome": tk.StringVar(),
            "straniero_indirizzo": tk.StringVar(),
            "data": tk.StringVar(value=_text((record_to_edit or {}).get("data_presentazione", ""))),
            "motivo": tk.StringVar(value="OSPITA"),
            "indirizzo_ospitalita": tk.StringVar(value=_text((record_to_edit or {}).get("indirizzo", ""))),
        }

        if record_to_edit:
            den_nome = _text(record_to_edit.get("denunciante_nome", "")).strip()
            den_ind = _text(record_to_edit.get("denunciante_indirizzo", "")).strip()
            if not den_nome and not den_ind:
                den_nome, den_ind = self._split_display_name_address(record_to_edit.get("denunciante_dichiarante", ""))
            st_nome = _text(record_to_edit.get("straniero_nome", "")).strip()
            st_ind = _text(record_to_edit.get("straniero_indirizzo", "")).strip()
            if not st_nome and not st_ind:
                st_nome, st_ind = self._split_display_name_address(record_to_edit.get("cittadino_ospitato", ""))
            motivo_value = _text(record_to_edit.get("tipo_comunicazione", "")).strip().upper() or "OSPITA"
            if motivo_value not in ("OSPITA", "CESSIONE IMMOBILE"):
                motivo_value = "OSPITA"
            vars_map["denunciante_nome"].set(den_nome)
            vars_map["denunciante_indirizzo"].set(den_ind)
            vars_map["straniero_nome"].set(st_nome)
            vars_map["straniero_indirizzo"].set(st_ind)
            vars_map["motivo"].set(motivo_value)

        fields = [
            ("Numero protocollo", "protocollo"),
            ("Cognome e Nome del denunciante", "denunciante_nome"),
            ("Indirizzo di residenza del denunciante", "denunciante_indirizzo"),
            ("Cognome e Nome dello straniero", "straniero_nome"),
            ("Indirizzo di residenza dello straniero", "straniero_indirizzo"),
            ("Data (GG/MM/AAAA)", "data"),
        ]

        for label, key in fields:
            row = tk.Frame(body, bg=SURFACE)
            row.pack(fill="x", padx=12, pady=6)
            tk.Label(
                row,
                text=label,
                bg=SURFACE,
                fg=TEXT_MUTED,
                font=("Segoe UI", 9, "bold"),
                width=42,
                anchor="w",
            ).pack(side="left")
            tk.Entry(
                row,
                textvariable=vars_map[key],
                font=("Segoe UI", 10),
                bg="white",
                fg=TEXT,
                relief="solid",
                bd=1,
            ).pack(side="left", fill="x", expand=True, ipady=4)

        motivo_row = tk.Frame(body, bg=SURFACE)
        motivo_row.pack(fill="x", padx=12, pady=6)
        tk.Label(
            motivo_row,
            text="Ospita o Cessione immobile",
            bg=SURFACE,
            fg=TEXT_MUTED,
            font=("Segoe UI", 9, "bold"),
            width=42,
            anchor="w",
        ).pack(side="left")
        motivo_combo = ttk.Combobox(
            motivo_row,
            textvariable=vars_map["motivo"],
            values=("OSPITA", "CESSIONE IMMOBILE"),
            state="readonly",
            font=("Segoe UI", 10),
            width=30,
        )
        motivo_combo.pack(side="left", fill="x", expand=True)

        indirizzo_row = tk.Frame(body, bg=SURFACE)
        indirizzo_row.pack(fill="x", padx=12, pady=6)
        tk.Label(
            indirizzo_row,
            text="Indirizzo (ospitalita/immobile)",
            bg=SURFACE,
            fg=TEXT_MUTED,
            font=("Segoe UI", 9, "bold"),
            width=42,
            anchor="w",
        ).pack(side="left")
        tk.Entry(
            indirizzo_row,
            textvariable=vars_map["indirizzo_ospitalita"],
            font=("Segoe UI", 10),
            bg="white",
            fg=TEXT,
            relief="solid",
            bd=1,
        ).pack(side="left", fill="x", expand=True, ipady=4)

        footer = tk.Frame(body, bg=SURFACE)
        footer.pack(fill="x", padx=12, pady=(8, 12))

        def conferma():
            protocollo = vars_map["protocollo"].get().strip()
            denunciante_nome = vars_map["denunciante_nome"].get().strip()
            denunciante_indirizzo = vars_map["denunciante_indirizzo"].get().strip()
            straniero_nome = vars_map["straniero_nome"].get().strip()
            straniero_indirizzo = vars_map["straniero_indirizzo"].get().strip()
            data = vars_map["data"].get().strip()
            motivo = vars_map["motivo"].get().strip().upper()
            indirizzo_ospitalita = vars_map["indirizzo_ospitalita"].get().strip()

            if not protocollo:
                messagebox.showwarning("Dati incompleti", "Inserisci il protocollo.", parent=win)
                return
            if not denunciante_nome:
                messagebox.showwarning("Dati incompleti", "Inserisci il cognome e nome del denunciante.", parent=win)
                return
            if not denunciante_indirizzo:
                messagebox.showwarning("Dati incompleti", "Inserisci l'indirizzo del denunciante.", parent=win)
                return
            if not straniero_nome:
                messagebox.showwarning("Dati incompleti", "Inserisci il cognome e nome dello straniero.", parent=win)
                return
            if not straniero_indirizzo:
                messagebox.showwarning("Dati incompleti", "Inserisci l'indirizzo dello straniero.", parent=win)
                return
            if _parse_sort_date(data) is None:
                messagebox.showwarning("Data non valida", "Data non valida. Usa GG/MM/AAAA.", parent=win)
                return
            if motivo not in ("OSPITA", "CESSIONE IMMOBILE"):
                messagebox.showwarning("Valore non valido", "Seleziona OSPITA o CESSIONE IMMOBILE.", parent=win)
                return
            if not indirizzo_ospitalita:
                messagebox.showwarning("Dati incompleti", "Inserisci l'indirizzo nel dettaglio ospitalita.", parent=win)
                return

            if not messagebox.askyesno(
                "Conferma inserimento",
                "Vuoi salvare questo inserimento nella copia di lavoro?",
                parent=win,
            ):
                return

            if is_edit and record_to_edit is not None:
                progressivo = str(record_to_edit.get("progressivo", "")).strip()
                source_name = str(record_to_edit.get("source", "")).strip() or self._current_source_name()
                target_record = record_to_edit
                original_snapshot = dict(record_to_edit)
                mode = "update_existing"
            else:
                progressivo, target_record, original_snapshot, mode = self._reserve_progressivo_slot()
                source_name = self._current_source_name()
            denunciante_display = self._compact_values(denunciante_nome, denunciante_indirizzo)
            straniero_display = self._compact_values(straniero_nome, straniero_indirizzo)

            if target_record is None:
                target_record = {
                    "progressivo": progressivo,
                    "source": source_name,
                    "sheet": "",
                    "_pending": True,
                }
                self.all_records.append(target_record)
            else:
                target_record["_pending"] = True

            target_record.update({
                "progressivo": progressivo,
                "source": source_name,
                "protocollo": protocollo,
                "denunciante_dichiarante": denunciante_display,
                "cittadino_ospitato": straniero_display,
                "data_presentazione": data,
                "indirizzo": indirizzo_ospitalita,
                "tipo_comunicazione": motivo,
                "denunciante_nome": denunciante_nome,
                "denunciante_indirizzo": denunciante_indirizzo,
                "straniero_nome": straniero_nome,
                "straniero_indirizzo": straniero_indirizzo,
                "dettaglio_motivo": motivo,
                "dettaglio_indirizzo": indirizzo_ospitalita,
            })

            self._upsert_pending_record({
                "mode": mode,
                "progressivo": progressivo,
                "source": source_name,
                "protocollo": protocollo,
                "data": data,
                "denunciante_nome": denunciante_nome,
                "denunciante_indirizzo": denunciante_indirizzo,
                "straniero_nome": straniero_nome,
                "straniero_indirizzo": straniero_indirizzo,
                "motivo": motivo,
                "indirizzo_ospitalita": indirizzo_ospitalita,
                "original_snapshot": original_snapshot,
            })
            self.btn_save_changes.config(state="normal")
            self.applica_filtro()
            win.destroy()
            messagebox.showinfo(
                "Modifica registrata" if is_edit else "Inserimento registrato",
                (
                    "Nominativo aggiornato nella copia di lavoro.\nUsa 'SALVA MODIFICHE' per aggiornare il file Excel."
                    if is_edit
                    else "Nominativo aggiunto nella copia di lavoro.\nUsa 'SALVA MODIFICHE' per aggiornare il file Excel."
                ),
            )

        tk.Button(
            footer,
            text="Salva inserimento",
            bg=ACCENT,
            fg="white",
            font=("Segoe UI", 10, "bold"),
            relief="flat",
            cursor="hand2",
            activebackground=ACCENT_DARK,
            padx=14,
            pady=8,
            command=conferma,
        ).pack(side="left")
        tk.Button(
            footer,
            text="Annulla",
            bg=BG2,
            fg=TEXT_MUTED,
            font=("Segoe UI", 10, "bold"),
            relief="flat",
            cursor="hand2",
            activebackground=BORDER,
            highlightbackground=BORDER,
            highlightthickness=1,
            padx=14,
            pady=8,
            command=win.destroy,
        ).pack(side="left", padx=(8, 0))

    def _append_pending_with_excel_com(self, workbook_path: Path):
        payload = [
            {
                "progressivo": str(rec.get("progressivo", "")),
                "protocollo": str(rec.get("protocollo", "")),
                "data": str(rec.get("data", "")),
                "denunciante_nome": str(rec.get("denunciante_nome", "")),
                "denunciante_indirizzo": str(rec.get("denunciante_indirizzo", "")),
                "straniero_nome": str(rec.get("straniero_nome", "")),
                "straniero_indirizzo": str(rec.get("straniero_indirizzo", "")),
                "motivo": str(rec.get("motivo", "")),
                "indirizzo_ospitalita": str(rec.get("indirizzo_ospitalita", "")),
            }
            for rec in self._pending_new_records
        ]

        ps_script = r"""
param(
    [Parameter(Mandatory = $true)][string]$WorkbookPath,
    [Parameter(Mandatory = $true)][string]$PayloadPath
)
$ErrorActionPreference = "Stop"
$items = Get-Content -Path $PayloadPath -Raw -Encoding UTF8 | ConvertFrom-Json
$excel = $null
$wb = $null
$ws = $null
try {
    $excel = New-Object -ComObject Excel.Application
    $excel.Visible = $false
    $excel.DisplayAlerts = $false
    $wb = $excel.Workbooks.Open($WorkbookPath)
    $ws = $wb.Worksheets.Item(1)
    foreach ($item in $items) {
        $targetTop = $null
        $prog = [string]$item.progressivo

        $usedStart = $ws.UsedRange.Row
        $usedEnd = $ws.UsedRange.Row + $ws.UsedRange.Rows.Count - 1
        for ($r = $usedStart; $r -le $usedEnd; $r++) {
            $value = [string]$ws.Cells.Item($r, 1).Text
            if ([string]::IsNullOrWhiteSpace($value)) { continue }
            if ($value.Trim() -eq $prog.Trim()) {
                $targetTop = $r
                break
            }
        }

        if ($targetTop -eq $null) {
            $targetTop = $usedEnd + 1
            $ws.Cells.Item($targetTop, 1).Value2 = $prog
        }

        $mergeRows = 1
        try {
            if ($ws.Cells.Item($targetTop, 1).MergeCells) {
                $mergeRows = $ws.Cells.Item($targetTop, 1).MergeArea.Rows.Count
            }
        } catch {}
        if ($mergeRows -lt 2) { $mergeRows = 2 }
        $targetBottom = $targetTop + $mergeRows - 1

        $ws.Cells.Item($targetTop, 2).Value2 = [string]$item.protocollo
        $ws.Cells.Item($targetTop, 3).Value2 = [string]$item.data
        $ws.Cells.Item($targetTop, 4).Value2 = [string]$item.denunciante_nome
        $ws.Cells.Item($targetBottom, 4).Value2 = [string]$item.denunciante_indirizzo
        $ws.Cells.Item($targetTop, 5).Value2 = [string]$item.straniero_nome
        $ws.Cells.Item($targetBottom, 5).Value2 = [string]$item.straniero_indirizzo
        $ws.Cells.Item($targetTop, 6).Value2 = [string]$item.motivo
        $ws.Cells.Item($targetBottom, 6).Value2 = [string]$item.indirizzo_ospitalita
    }

    $wb.Save()
}
finally {
    if ($wb -ne $null) { $wb.Close($true) }
    if ($excel -ne $null) { $excel.Quit() }
}
"""
        with tempfile.TemporaryDirectory(prefix="passapp_ospitalita_") as tmp_dir:
            tmp_path = Path(tmp_dir)
            payload_path = tmp_path / "rows.json"
            script_path = tmp_path / "append.ps1"
            payload_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
            script_path.write_text(ps_script, encoding="utf-8")
            result = subprocess.run(
                [
                    "powershell",
                    "-NoProfile",
                    "-ExecutionPolicy",
                    "Bypass",
                    "-File",
                    str(script_path),
                    "-WorkbookPath",
                    str(workbook_path),
                    "-PayloadPath",
                    str(payload_path),
                ],
                capture_output=True,
                text=True,
                encoding="utf-8",
                errors="replace",
                timeout=180,
            )
            if result.returncode != 0:
                details = (result.stderr or result.stdout or "Errore sconosciuto").strip()
                raise RuntimeError(details)

    def salva_modifiche(self, trigger_reload: bool = True) -> bool:
        if not self._pending_new_records:
            messagebox.showinfo("Nessuna modifica", "Non ci sono modifiche da salvare.")
            return False
        if not self._working_copy_file or not self._primary_source_file:
            messagebox.showwarning("Salvataggio non disponibile", "Copia di lavoro non disponibile.")
            return False

        suffix = self._working_copy_file.suffix.lower()
        if suffix not in (".xlsx", ".xls"):
            messagebox.showwarning(
                "Formato non supportato",
                "Il salvataggio modifiche e supportato per file .xls/.xlsx nel modulo Ospitalita.",
            )
            return False

        try:
            self._append_pending_with_excel_com(self._working_copy_file)
            shutil.copy2(self._working_copy_file, self._primary_source_file)
        except Exception as exc:
            logger.exception("Errore salvataggio modifiche ospitalita")
            messagebox.showerror("Salvataggio non riuscito", f"Impossibile salvare le modifiche.\n\nDettagli:\n{exc}")
            return False

        for rec in self.all_records:
            rec.pop("_pending", None)
        self._pending_new_records = []
        self.btn_save_changes.config(state="disabled")
        if trigger_reload:
            self.carica_dati(force=True)
        else:
            self.applica_filtro()
        messagebox.showinfo("Salvataggio completato", "Le modifiche sono state salvate sul file Excel.")
        return True

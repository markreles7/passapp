from __future__ import annotations

import datetime as dt
import json
import os
import shutil
import tempfile
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from app_config import resolve_path
from core.dates import parse_date
from core.logging_utils import setup_module_logger

ACCERTAMENTI_FILE = resolve_path("data/accertamenti_anagrafici.json")
ACCERTAMENTI_BACKUP_FILE = Path(f"{ACCERTAMENTI_FILE}.bak")
ACCERTAMENTI_MALFORMED_BACKUP_DIR = resolve_path("data/backups/accertamenti_anagrafici")
DEFAULT_ACCERTAMENTI_EXCEL = Path(r"R:\Polizia_locale\ACCERTAMENTI ANAGRAFICI.xlsx")
logger = setup_module_logger(__name__, resolve_path("data/passapp.log"))

STATI_ACCERTAMENTO = (
    "da fare",
    "in corso",
    "completato",
)


@dataclass
class AccertamentoAnagrafico:
    numero: int
    nominativo: str = ""
    indirizzo: str = ""
    note: str = ""
    primo_negativo_data: str = ""
    primo_negativo_ora: str = ""
    secondo_negativo_data: str = ""
    secondo_negativo_ora: str = ""
    terzo_negativo_data: str = ""
    terzo_negativo_ora: str = ""
    positivo_data: str = ""
    positivo_ora: str = ""
    data_creazione: str = ""
    data_ultima_modifica: str = ""

    @classmethod
    def from_dict(cls, raw: dict[str, Any]) -> AccertamentoAnagrafico | None:
        try:
            numero = int(raw.get("numero", 0))
        except (TypeError, ValueError):
            return None
        if numero <= 0:
            return None
        return cls(
            numero=numero,
            nominativo=str(raw.get("nominativo", "")),
            indirizzo=str(raw.get("indirizzo", "")),
            note=str(raw.get("note", "")),
            primo_negativo_data=str(raw.get("primo_negativo_data", "")),
            primo_negativo_ora=str(raw.get("primo_negativo_ora", "")),
            secondo_negativo_data=str(raw.get("secondo_negativo_data", "")),
            secondo_negativo_ora=str(raw.get("secondo_negativo_ora", "")),
            terzo_negativo_data=str(raw.get("terzo_negativo_data", "")),
            terzo_negativo_ora=str(raw.get("terzo_negativo_ora", "")),
            positivo_data=str(raw.get("positivo_data", "")),
            positivo_ora=str(raw.get("positivo_ora", "")),
            data_creazione=str(raw.get("data_creazione", "")),
            data_ultima_modifica=str(raw.get("data_ultima_modifica", "")),
        )

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)

    @property
    def stato(self) -> str:
        return stato_accertamento(self)


def load_accertamenti(path: Path = ACCERTAMENTI_FILE) -> list[AccertamentoAnagrafico]:
    if not path.exists():
        return []
    try:
        with path.open("r", encoding="utf-8") as handle:
            payload = json.load(handle)
    except json.JSONDecodeError:
        backup = backup_malformed_accertamenti_file(path)
        logger.exception("File accertamenti JSON malformato. Backup creato: %s", backup)
        return []
    except OSError:
        logger.exception("File accertamenti non leggibile: %s", path)
        return []
    items = payload.get("accertamenti", []) if isinstance(payload, dict) else payload
    if not isinstance(items, list):
        return []
    out: list[AccertamentoAnagrafico] = []
    for item in items:
        if not isinstance(item, dict):
            continue
        accertamento = AccertamentoAnagrafico.from_dict(item)
        if accertamento is not None:
            out.append(accertamento)
    return sorted(out, key=lambda item: item.numero)


def save_accertamenti(
    accertamenti: list[AccertamentoAnagrafico],
    path: Path = ACCERTAMENTI_FILE,
    backup_path: Path | None = None,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    backup_target = backup_path if backup_path is not None else Path(f"{path}.bak")
    fd, tmp_name = tempfile.mkstemp(prefix="accertamenti_", suffix=".tmp", dir=str(path.parent))
    tmp_path = Path(tmp_name)
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as handle:
            json.dump({"accertamenti": [item.to_dict() for item in accertamenti]}, handle, ensure_ascii=False, indent=2)
            handle.flush()
            os.fsync(handle.fileno())
        if path.exists():
            backup_target.write_bytes(path.read_bytes())
        os.replace(tmp_path, path)
    finally:
        if tmp_path.exists():
            try:
                tmp_path.unlink()
            except OSError:
                pass


def next_accertamento_numero(accertamenti: list[AccertamentoAnagrafico]) -> int:
    return max((item.numero for item in accertamenti), default=0) + 1


def upsert_accertamento(item: AccertamentoAnagrafico, path: Path = ACCERTAMENTI_FILE) -> None:
    records = load_accertamenti(path)
    found = False
    for idx, current in enumerate(records):
        if current.numero == item.numero:
            records[idx] = item
            found = True
            break
    if not found:
        records.append(item)
    save_accertamenti(records, path)


def delete_accertamento(numero: int, path: Path = ACCERTAMENTI_FILE) -> None:
    records = [item for item in load_accertamenti(path) if item.numero != numero]
    save_accertamenti(records, path)


def stato_accertamento(item: AccertamentoAnagrafico) -> str:
    if item.positivo_data.strip() and item.positivo_ora.strip():
        return "completato"
    for data, ora in negative_attempts(item):
        if data.strip() or ora.strip():
            return "in corso"
    return "da fare"


def negative_attempts(item: AccertamentoAnagrafico) -> tuple[tuple[str, str], tuple[str, str], tuple[str, str]]:
    return (
        (item.primo_negativo_data, item.primo_negativo_ora),
        (item.secondo_negativo_data, item.secondo_negativo_ora),
        (item.terzo_negativo_data, item.terzo_negativo_ora),
    )


def validate_accertamento(item: AccertamentoAnagrafico) -> tuple[bool, str]:
    if item.numero <= 0:
        return False, "Il numero progressivo deve essere maggiore di zero."
    if not item.nominativo.strip():
        return False, "Il nominativo e obbligatorio."
    if not item.indirizzo.strip():
        return False, "L'indirizzo e obbligatorio."
    checks = (
        ("1° negativo", item.primo_negativo_data, item.primo_negativo_ora),
        ("2° negativo", item.secondo_negativo_data, item.secondo_negativo_ora),
        ("3° negativo", item.terzo_negativo_data, item.terzo_negativo_ora),
        ("positivo", item.positivo_data, item.positivo_ora),
    )
    for label, data, ora in checks:
        ok, reason = validate_date_time_pair(label, data, ora)
        if not ok:
            return False, reason
    return True, ""


def validate_date_time_pair(label: str, data: str, ora: str) -> tuple[bool, str]:
    data = data.strip()
    ora = ora.strip()
    if not data and not ora:
        return True, ""
    if not data or not ora:
        return False, f"Per {label} inserisci sia la data sia l'ora."
    if parse_date(data) is None:
        return False, f"La data per {label} non e valida. Usa GG/MM/AAAA."
    if normalize_time(ora) is None:
        return False, f"L'ora per {label} non e valida. Usa HH:MM."
    return True, ""


def normalize_date(value: Any) -> str:
    parsed = parse_date(value)
    if parsed is not None:
        return parsed.strftime("%d/%m/%Y")
    return str(value).strip() if value is not None else ""


def normalize_time(value: Any) -> str | None:
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        return value.strftime("%H:%M")
    if isinstance(value, dt.time):
        return value.strftime("%H:%M")
    text = str(value).strip().replace(";", ":").replace(".", ":")
    if not text:
        return ""
    for fmt in ("%H:%M", "%H:%M:%S", "%H"):
        try:
            return dt.datetime.strptime(text, fmt).strftime("%H:%M")
        except ValueError:
            pass
    return None


def import_accertamenti_from_excel(path: Path = DEFAULT_ACCERTAMENTI_EXCEL) -> list[AccertamentoAnagrafico]:
    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active
    records: list[AccertamentoAnagrafico] = []
    for row in sheet.iter_rows(min_row=3, values_only=True):
        if not row or not any(value not in (None, "") for value in row[:10]):
            continue
        try:
            numero = int(row[0])
        except (TypeError, ValueError):
            numero = len(records) + 1
        record = AccertamentoAnagrafico(
            numero=numero,
            nominativo=_cell_text(row, 1),
            indirizzo=_cell_text(row, 2),
            note=_cell_text(row, 3),
            primo_negativo_data=normalize_date(_cell_value(row, 4)),
            primo_negativo_ora=normalize_time(_cell_value(row, 5)) or _cell_text(row, 5),
            secondo_negativo_data=normalize_date(_cell_value(row, 6)),
            secondo_negativo_ora=normalize_time(_cell_value(row, 7)) or _cell_text(row, 7),
            terzo_negativo_data=normalize_date(_cell_value(row, 8)),
            terzo_negativo_ora=normalize_time(_cell_value(row, 9)) or _cell_text(row, 9),
            positivo_data=normalize_date(_cell_value(row, 10)),
            positivo_ora=normalize_time(_cell_value(row, 11)) or _cell_text(row, 11),
            data_creazione=now_timestamp(),
            data_ultima_modifica=now_timestamp(),
        )
        if record.nominativo.strip() or record.indirizzo.strip():
            records.append(record)
    return records


def save_accertamenti_to_excel(path: Path, records: list[AccertamentoAnagrafico]) -> None:
    if path.exists():
        workbook = load_workbook(path)
    else:
        workbook = Workbook()
    sheet = workbook.active

    _ensure_excel_headers(sheet)
    max_row = max(sheet.max_row, len(records) + 2)
    if max_row >= 3:
        for row in sheet.iter_rows(min_row=3, max_row=max_row, min_col=1, max_col=12):
            for cell in row:
                cell.value = None

    for offset, record in enumerate(sorted(records, key=lambda item: item.numero), start=3):
        values = (
            record.numero,
            record.nominativo,
            record.indirizzo,
            record.note,
            record.primo_negativo_data,
            record.primo_negativo_ora,
            record.secondo_negativo_data,
            record.secondo_negativo_ora,
            record.terzo_negativo_data,
            record.terzo_negativo_ora,
            record.positivo_data,
            record.positivo_ora,
        )
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row=offset, column=column)
            cell.value = value
            if record.stato == "completato":
                cell.font = Font(color="808080", strike=True)
                cell.fill = PatternFill("solid", fgColor="F2F2F2")
            else:
                cell.font = Font(color="000000", strike=False)
                cell.fill = PatternFill(fill_type=None)

    _set_default_widths(sheet)
    workbook.save(path)


def initialize_accertamenti_from_excel_if_needed(
    json_path: Path = ACCERTAMENTI_FILE,
    excel_path: Path = DEFAULT_ACCERTAMENTI_EXCEL,
) -> bool:
    if json_path.exists() or not excel_path.exists():
        return False
    records = import_accertamenti_from_excel(excel_path)
    if not records:
        return False
    save_accertamenti(records, json_path)
    return True


def _ensure_excel_headers(sheet) -> None:
    for merged_range in list(sheet.merged_cells.ranges):
        if merged_range.min_row <= 2 and merged_range.max_row >= 1 and merged_range.min_col <= 12 and merged_range.max_col >= 1:
            sheet.unmerge_cells(str(merged_range))
    headers_top = (
        "N°",
        "NOMINATIVO",
        "INDIRIZZO",
        "NOTE",
        "1° Negativo",
        "",
        "2° Negativo",
        "",
        "3° Negativo",
        "",
        "POSITIVO",
        "",
    )
    headers_bottom = ("", "", "", "", "Data", "Ora", "Data", "Ora", "Data", "Ora", "Data", "Ora")
    for column, value in enumerate(headers_top, start=1):
        sheet.cell(row=1, column=column).value = value
        sheet.cell(row=1, column=column).font = Font(bold=True)
    for column, value in enumerate(headers_bottom, start=1):
        sheet.cell(row=2, column=column).value = value
        sheet.cell(row=2, column=column).font = Font(bold=True)
    for cell_range in ("A1:A2", "B1:B2", "C1:C2", "D1:D2", "E1:F1", "G1:H1", "I1:J1", "K1:L1"):
        sheet.merge_cells(cell_range)


def _set_default_widths(sheet) -> None:
    widths = {
        "A": 8,
        "B": 26,
        "C": 34,
        "D": 34,
        "E": 14,
        "F": 10,
        "G": 14,
        "H": 10,
        "I": 14,
        "J": 10,
        "K": 14,
        "L": 10,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def now_timestamp() -> str:
    return dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def backup_malformed_accertamenti_file(path: Path) -> Path:
    backup_dir = _backup_dir_for(path)
    backup_dir.mkdir(parents=True, exist_ok=True)
    backup = backup_dir / f"{path.stem}_malformed_{dt.datetime.now().strftime('%Y%m%d_%H%M%S')}{path.suffix}"
    shutil.copy2(path, backup)
    return backup


def _cell_value(row: tuple[Any, ...], index: int) -> Any:
    return row[index] if len(row) > index else None


def _cell_text(row: tuple[Any, ...], index: int) -> str:
    value = _cell_value(row, index)
    if value is None:
        return ""
    return str(value).strip()


def _backup_dir_for(path: Path) -> Path:
    try:
        if path.resolve() == ACCERTAMENTI_FILE.resolve():
            return ACCERTAMENTI_MALFORMED_BACKUP_DIR
    except OSError:
        pass
    return path.parent / "backups" / "accertamenti_anagrafici"

from __future__ import annotations

import difflib
import glob
import os
import re

from app_config import load_config, resolve_path
from core.dates import parse_date as _common_parse_date
from core.logging_utils import setup_module_logger
from core.text_utils import display_text, normalize_basic

try:
    import xlrd  # type: ignore
except ImportError:  # pragma: no cover
    xlrd = None

try:
    import openpyxl  # type: ignore
except ImportError:  # pragma: no cover
    openpyxl = None

APP_CONFIG = load_config()
PATHS = APP_CONFIG["paths"]

FOLDER_OSPITALITA = PATHS["ospitalita_network_folder"]
FILE_PATTERNS = list(PATHS["ospitalita_patterns"])
WORK_COPY_DIR = resolve_path("data/workcopies/ospitalita")
logger = setup_module_logger(__name__, resolve_path(PATHS["log_file"]))

FIELD_ALIASES = {
    "protocollo": ["protocollo", "n protocollo", "num protocollo", "prot"],
    "data_presentazione": ["data presentazione", "data comunicazione", "data denuncia", "data", "presentazione"],
    "denunciante_dichiarante": ["denunciante", "dichiarante", "comunicante", "richiedente", "segnalante"],
    "cittadino_ospitato": [
        "ospitato",
        "cittadino ospitato",
        "straniero ospitato",
        "ospite",
        "generalita ospitato",
        "cognome",
        "nome",
        "nominativo",
    ],
    "indirizzo": ["indirizzo", "via", "residenza", "domicilio", "alloggio", "luogo ospitalita"],
    "tipo_comunicazione": ["tipo comunicazione", "tipologia", "tipo", "comunicazione", "motivo"],
}


def _norm(value) -> str:
    return normalize_basic(value)


def _text(value) -> str:
    return display_text(value)


def _parse_sort_date(value):
    return _common_parse_date(value)


def _protocol_sort_key(value):
    text = _text(value)
    if not text:
        return (1, tuple(), "")
    nums = re.findall(r"\d+", text)
    if nums:
        return (0, tuple(int(num) for num in nums), _norm(text))
    return (1, tuple(), _norm(text))


def _extract_year_from_name(name: str) -> int:
    years = [int(match.group(0)) for match in re.finditer(r"(?:19|20)\d{2}", name)]
    return max(years) if years else -1


def _file_sort_key(path: str):
    base = os.path.basename(path)
    year = _extract_year_from_name(base)
    try:
        mtime = os.path.getmtime(path)
    except OSError:
        mtime = 0.0
    return (year, mtime, base.lower())


def _looks_empty_row(row) -> bool:
    return all(_norm(cell) == "" for cell in row)


def _tokens(text: str) -> set[str]:
    return {token for token in re.split(r"[^a-z0-9]+", text) if token}


def _match_score(header_text: str, alias: str) -> float:
    header = _norm(header_text)
    alias_norm = _norm(alias)
    if not header or not alias_norm:
        return 0.0
    if header == alias_norm:
        return 1.0
    if alias_norm in header:
        return 0.92
    if header in alias_norm:
        return 0.76

    header_tokens = _tokens(header)
    alias_tokens = _tokens(alias_norm)
    overlap = 0.0
    if header_tokens and alias_tokens:
        overlap = len(header_tokens & alias_tokens) / max(len(alias_tokens), 1)

    ratio = difflib.SequenceMatcher(None, header, alias_norm).ratio()
    return max(overlap * 0.9, ratio * 0.8)


def _best_alias_score(header_text: str, aliases: list[str]) -> float:
    return max((_match_score(header_text, alias) for alias in aliases), default=0.0)


def _header_score(row) -> float:
    headers = [_norm(cell) for cell in row if _norm(cell)]
    if not headers:
        return 0.0
    score = 0.0
    for head in headers:
        score += max((_best_alias_score(head, aliases) for aliases in FIELD_ALIASES.values()), default=0.0)
    return score


def _find_header_row(rows) -> int:
    best_idx = 0
    best_score = -1.0
    for idx in range(min(30, len(rows))):
        score = _header_score(rows[idx])
        if score > best_score:
            best_idx = idx
            best_score = score
    return best_idx


def _map_columns(headers) -> dict[str, int | None]:
    mapping: dict[str, int | None] = {}
    used: set[int] = set()
    for field, aliases in FIELD_ALIASES.items():
        best_idx = None
        best_score = 0.0
        for idx, head in enumerate(headers):
            if idx in used:
                continue
            score = _best_alias_score(head, aliases)
            if score > best_score:
                best_score = score
                best_idx = idx
        if best_idx is not None and best_score >= 0.45:
            mapping[field] = best_idx
            used.add(best_idx)
        else:
            mapping[field] = None
    return mapping


def _load_rows_xls(path: str):
    if xlrd is None:
        logger.error("Import .xls non possibile: libreria xlrd non disponibile (%s)", path)
        raise RuntimeError("Libreria 'xlrd' non disponibile per leggere file .xls")
    book = xlrd.open_workbook(path)
    best_sheet_idx = 0
    best_sheet_score = -1.0
    best_header_idx = 0
    for idx in range(book.nsheets):
        sheet = book.sheet_by_index(idx)
        sample = [[sheet.cell_value(row, col) for col in range(sheet.ncols)] for row in range(min(sheet.nrows, 35))]
        if not sample:
            continue
        header_idx = _find_header_row(sample)
        score = _header_score(sample[header_idx]) if header_idx < len(sample) else 0.0
        if score > best_sheet_score:
            best_sheet_score = score
            best_sheet_idx = idx
            best_header_idx = header_idx

    sheet = book.sheet_by_index(best_sheet_idx)
    rows = []
    for row in range(sheet.nrows):
        parsed_row = []
        for col in range(sheet.ncols):
            cell = sheet.cell(row, col)
            value = cell.value
            if cell.ctype == getattr(xlrd, "XL_CELL_DATE", 3):
                try:
                    value = xlrd.xldate_as_datetime(value, book.datemode)
                except Exception:
                    pass
            parsed_row.append(value)
        rows.append(parsed_row)
    return rows, {"active_sheet": sheet.name, "active_sheet_index": best_sheet_idx, "header_hint": best_header_idx + 1}


def _load_rows_xlsx(path: str):
    if openpyxl is None:
        logger.error("Import .xlsx non possibile: libreria openpyxl non disponibile (%s)", path)
        raise RuntimeError("Libreria 'openpyxl' non disponibile per leggere file .xlsx")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    active_title = wb.active.title
    sheet_rows = [(ws.title, [list(row) for row in ws.iter_rows(values_only=True)]) for ws in wb.worksheets]

    selected_title = active_title
    selected_rows = []
    best_score = -1.0
    for title, rows in sheet_rows:
        if not rows:
            continue
        header_idx = _find_header_row(rows)
        score = _header_score(rows[header_idx]) if header_idx < len(rows) else 0.0
        if title == active_title:
            selected_rows = rows
            best_score = score
        if score > best_score + 0.25:
            selected_title = title
            selected_rows = rows
            best_score = score
    if not selected_rows:
        selected_rows = [list(row) for row in wb.active.iter_rows(values_only=True)]
    wb.close()
    return selected_rows, {"active_sheet": selected_title, "workbook_active_sheet": active_title}


def _load_rows(path: str):
    lower = path.lower()
    if lower.endswith(".xlsx"):
        return _load_rows_xlsx(path)
    if lower.endswith(".xls"):
        return _load_rows_xls(path)
    raise RuntimeError("Formato file non supportato")


def _extract_records(path: str):
    rows, meta = _load_rows(path)
    if not rows:
        return [], "Nessuna riga trovata"

    header_idx = _find_header_row(rows)
    headers = rows[header_idx]
    mapping = _map_columns(headers)
    records = []

    def pick(row, key):
        idx = mapping.get(key)
        if idx is None or idx >= len(row):
            return ""
        return _text(row[idx])

    def progressivo(row):
        value = _text(row[0] if row else "")
        marker = _norm(value)
        if marker in {"", "n", "n.", "nr", "numero", "num", "progressivo", "n°"}:
            return ""
        return value

    def merge_value(current: str, extra: str) -> str:
        cur = _text(current).strip()
        add = _text(extra).strip()
        if not add or _norm(add) in {"-", "—"}:
            return cur
        if not cur or _norm(cur) in {"-", "—"}:
            return add
        parts = [part.strip() for part in cur.split(" | ") if part.strip()]
        if add not in parts:
            parts.append(add)
        return " | ".join(parts)

    for row in rows[header_idx + 1 :]:
        if _looks_empty_row(row):
            continue
        record = {
            "progressivo": progressivo(row),
            "protocollo": pick(row, "protocollo"),
            "data_presentazione": pick(row, "data_presentazione"),
            "denunciante_dichiarante": pick(row, "denunciante_dichiarante"),
            "cittadino_ospitato": pick(row, "cittadino_ospitato"),
            "indirizzo": pick(row, "indirizzo"),
            "tipo_comunicazione": pick(row, "tipo_comunicazione"),
            "source": os.path.basename(path),
            "sheet": meta.get("active_sheet", ""),
        }
        has_any_data = any(
            record[key]
            for key in (
                "progressivo",
                "protocollo",
                "data_presentazione",
                "denunciante_dichiarante",
                "cittadino_ospitato",
                "indirizzo",
                "tipo_comunicazione",
            )
        )
        if not has_any_data:
            continue
        is_continuation = not (record["progressivo"] or record["protocollo"] or record["data_presentazione"])
        if is_continuation and records:
            prev = records[-1]
            for key in ("denunciante_dichiarante", "cittadino_ospitato", "indirizzo", "tipo_comunicazione"):
                prev[key] = merge_value(prev.get(key, ""), record.get(key, ""))
            continue
        records.append(record)

    mapped_cols = ", ".join(f"{key}:{value}" for key, value in mapping.items())
    active_info = meta.get("active_sheet", "")
    workbook_active = meta.get("workbook_active_sheet")
    if workbook_active and workbook_active != active_info:
        sheet_info = f"sheet={active_info} (active workbook={workbook_active})"
    else:
        sheet_info = f"sheet={active_info}"
    insight = f"{os.path.basename(path)} -> {sheet_info} header row {header_idx + 1} ({mapped_cols})"
    return records, insight


def _list_input_files() -> list[str]:
    files: list[str] = []
    for pattern in FILE_PATTERNS:
        files.extend(glob.glob(os.path.join(FOLDER_OSPITALITA, pattern)))
    return sorted(set(files), key=_file_sort_key, reverse=True)


def _detect_layout(path: str) -> dict:
    rows, meta = _load_rows(path)
    if not rows:
        return {"sheet": meta.get("active_sheet", ""), "mapping": {}, "header_idx": 0}
    header_idx = _find_header_row(rows)
    headers = rows[header_idx] if header_idx < len(rows) else []
    mapping = _map_columns(headers)
    return {"sheet": meta.get("active_sheet", ""), "mapping": mapping, "header_idx": header_idx}

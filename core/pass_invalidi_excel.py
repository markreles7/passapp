from __future__ import annotations

import datetime
import json
import re
import subprocess
import tempfile
from pathlib import Path

import openpyxl

from core.dates import parse_date


def split_cognome_nome(full_name: str) -> tuple[str, str]:
    tokens = [tok for tok in re.split(r"\s+", str(full_name or "").strip()) if tok]
    if not tokens:
        return "", ""
    cognome = tokens[0].upper()
    nome = " ".join(tokens[1:]) if len(tokens) > 1 else ""
    nome = " ".join(part.capitalize() for part in nome.split()) if nome else ""
    return cognome, nome


def compose_full_name(cognome: str, nome: str) -> str:
    left = str(cognome or "").strip().upper()
    right = str(nome or "").strip().upper()
    if left and right:
        return f"{left} {right}"
    return left or right


def compose_generalita_indirizzo(cognome: str, nome: str, indirizzo: str) -> str:
    full_name = compose_full_name(cognome, nome)
    address = str(indirizzo or "").strip().upper()
    if full_name and address:
        return f"{full_name} - {address}"
    return full_name or address


def compute_authorization_expiry(protocol_date: datetime.date, birth_date: datetime.date) -> datetime.date:
    plus_years = 5 if (protocol_date.month, protocol_date.day) <= (birth_date.month, birth_date.day) else 6
    target_year = protocol_date.year + plus_years
    day = birth_date.day
    month = birth_date.month
    while day > 0:
        try:
            return datetime.date(target_year, month, day)
        except ValueError:
            day -= 1
    return datetime.date(target_year, month, 1)


def record_numero_int(record: dict) -> int | None:
    try:
        return int(str(record.get("numero", "")).strip())
    except (TypeError, ValueError):
        return None


def is_empty_slot(record: dict) -> bool:
    fields = (
        record.get("nome", ""),
        record.get("indirizzo", ""),
        record.get("rilascio", ""),
        record.get("scadenza", ""),
        record.get("note", ""),
    )
    return all(str(value or "").strip().lower() in {"", "-", "none", "nan"} for value in fields)


def find_record_by_numero(records: list[dict], numero: int, source_name: str | None = None) -> dict | None:
    source_ref = (source_name or "").strip()
    for record in records:
        if source_ref and str(record.get("source", "")).strip() != source_ref:
            continue
        if record_numero_int(record) == numero:
            return record
    return None


def reserve_numero_slot(records: list[dict], source_name: str) -> tuple[int, dict | None, dict | None, str]:
    source_ref = source_name.strip()
    numeric_records: list[tuple[int, dict]] = []
    for record in records:
        if source_ref and str(record.get("source", "")).strip() != source_ref:
            continue
        numero = record_numero_int(record)
        if numero is not None:
            numeric_records.append((numero, record))

    for numero, record in sorted(numeric_records, key=lambda item: item[0]):
        if record.get("_pending"):
            continue
        if is_empty_slot(record):
            return numero, record, dict(record), "update_slot"

    max_numero = max((numero for numero, _ in numeric_records), default=0)
    return max_numero + 1, None, None, "append"


def upsert_pending_record(pending_records: list[dict], pending: dict) -> None:
    source = str(pending.get("source", "")).strip()
    numero = str(pending.get("numero", "")).strip()
    for idx, existing in enumerate(pending_records):
        if str(existing.get("source", "")).strip() == source and str(existing.get("numero", "")).strip() == numero:
            if existing.get("original_snapshot") and not pending.get("original_snapshot"):
                pending["original_snapshot"] = existing.get("original_snapshot")
            pending_records[idx] = pending
            return
    pending_records.append(pending)


def maybe_date_value(text):
    parsed = parse_date(text)
    return parsed if parsed is not None else text


def write_pending_to_xlsx(workbook_path: Path, pending_records: list[dict]) -> None:
    wb = openpyxl.load_workbook(workbook_path)
    ws = wb.active

    wb_values = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    ws_values = wb_values.active
    row_by_numero: dict[int, int] = {}
    empty_rows: list[int] = []
    next_append_row = ws.max_row + 1
    for row_idx in range(1, ws_values.max_row + 1):
        num_val = ws_values.cell(row=row_idx, column=1).value
        text_val = ws_values.cell(row=row_idx, column=2).value
        try:
            numero = int(float(str(num_val).strip()))
        except (TypeError, ValueError):
            numero = None
        if numero is not None:
            row_by_numero[numero] = row_idx
            if text_val is None or str(text_val).strip() == "":
                empty_rows.append(row_idx)
    wb_values.close()

    for record in pending_records:
        numero = record_numero_int(record)
        if numero is None:
            continue

        target_row = row_by_numero.get(numero)
        if record.get("mode") == "delete_existing":
            if target_row is not None:
                for column in range(1, 6):
                    ws.cell(row=target_row, column=column).value = None
            continue

        if target_row is None:
            if empty_rows:
                target_row = empty_rows.pop(0)
            else:
                target_row = next_append_row
                next_append_row += 1
            ws.cell(row=target_row, column=1, value=numero)
            row_by_numero[numero] = target_row
        elif target_row in empty_rows:
            empty_rows.remove(target_row)

        generalita = compose_generalita_indirizzo(
            record.get("cognome", ""),
            record.get("nome_proprio", ""),
            record.get("indirizzo", ""),
        )
        ws.cell(row=target_row, column=2, value=generalita)
        ws.cell(row=target_row, column=3, value=maybe_date_value(record.get("rilascio", "")))
        ws.cell(row=target_row, column=4, value=maybe_date_value(record.get("scadenza", "")))
        ws.cell(row=target_row, column=5, value=record.get("note", ""))

    wb.save(workbook_path)
    wb.close()


def _numero_da_cella(value) -> int | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        try:
            return int(value)
        except (TypeError, ValueError):
            return None
    text = str(value).strip()
    if not text:
        return None
    match = re.match(r"^(\d+)(?:\s|$)", text)
    if not match:
        return None
    try:
        return int(match.group(1))
    except (TypeError, ValueError):
        return None


def write_pending_with_excel_com(workbook_path: Path, pending_records: list[dict]) -> None:
    wb_values = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    ws_values = wb_values.active

    row_by_numero: dict[int, int] = {}
    empty_rows: list[int] = []
    last_numbered_row = 1
    for row_idx in range(1, ws_values.max_row + 1):
        numero = _numero_da_cella(ws_values.cell(row=row_idx, column=1).value)
        if numero is None:
            continue
        if numero not in row_by_numero:
            row_by_numero[numero] = row_idx
        if row_idx > last_numbered_row:
            last_numbered_row = row_idx
        generalita = ws_values.cell(row=row_idx, column=2).value
        if generalita is None or str(generalita).strip() == "":
            empty_rows.append(row_idx)
    wb_values.close()

    payload = [
        {
            "numero": record.get("numero", ""),
            "generalita": compose_generalita_indirizzo(
                record.get("cognome", ""),
                record.get("nome_proprio", ""),
                record.get("indirizzo", ""),
            ),
            "indirizzo": record.get("indirizzo", ""),
            "rilascio": record.get("rilascio", ""),
            "scadenza": record.get("scadenza", ""),
            "note": record.get("note", ""),
            "mode": record.get("mode", ""),
            "target_row": None,
            "set_numero": False,
        }
        for record in pending_records
    ]

    for item in payload:
        try:
            numero = int(str(item.get("numero", "")).strip())
        except (TypeError, ValueError):
            continue

        if item.get("mode") == "delete_existing":
            item["target_row"] = row_by_numero.get(numero)
            item["set_numero"] = False
            continue

        target_row = row_by_numero.get(numero)
        set_numero = False
        if target_row is None:
            if empty_rows:
                target_row = empty_rows.pop(0)
            else:
                last_numbered_row += 1
                target_row = last_numbered_row
                set_numero = True
        elif target_row in empty_rows:
            empty_rows.remove(target_row)
        item["target_row"] = target_row
        item["set_numero"] = set_numero
        row_by_numero[numero] = target_row

    ps_script = r"""
param(
    [Parameter(Mandatory = $true)][string]$WorkbookPath,
    [Parameter(Mandatory = $true)][string]$PayloadPath
)
$ErrorActionPreference = "Stop"
$items = Get-Content -Path $PayloadPath -Raw -Encoding UTF8 | ConvertFrom-Json
$excel = $null
$wb = $null
$ws = $null
try {
    $excel = New-Object -ComObject Excel.Application
    $excel.Visible = $false
    $excel.DisplayAlerts = $false
    $wb = $excel.Workbooks.Open($WorkbookPath)
    $ws = $wb.Worksheets.Item(1)
    foreach ($item in $items) {
        $targetRow = 0
        [void][int]::TryParse([string]$item.target_row, [ref]$targetRow)
        $setNumero = $false
        try { $setNumero = [bool]$item.set_numero } catch {}
        $numero = [string]$item.numero
        $usedStart = $ws.UsedRange.Row
        $usedEnd = $ws.UsedRange.Row + $ws.UsedRange.Rows.Count - 1

        if ($targetRow -le 0) {
            for ($r = $usedStart; $r -le $usedEnd; $r++) {
                $value = [string]$ws.Cells.Item($r, 1).Text
                if ([string]::IsNullOrWhiteSpace($value)) { continue }
                if ($value.Trim() -eq $numero.Trim()) {
                    $targetRow = $r
                    break
                }
            }
        }
        if ($targetRow -le 0) {
            for ($r = $usedStart; $r -le $usedEnd; $r++) {
                $numText = [string]$ws.Cells.Item($r, 1).Text
                $genText = [string]$ws.Cells.Item($r, 2).Text
                if (-not [string]::IsNullOrWhiteSpace($numText) -and [string]::IsNullOrWhiteSpace($genText)) {
                    $targetRow = $r
                    break
                }
            }
        }
        if ($targetRow -le 0) {
            $targetRow = $usedEnd + 1
            $setNumero = $true
        }

        if ([string]$item.mode -eq "delete_existing") {
            if ($targetRow -gt 0) {
                $ws.Range($ws.Cells.Item($targetRow, 1), $ws.Cells.Item($targetRow, 5)).ClearContents()
            }
            continue
        }

        if ($setNumero -or [string]::IsNullOrWhiteSpace([string]$ws.Cells.Item($targetRow, 1).Text)) {
            $ws.Cells.Item($targetRow, 1).Value2 = [string]$item.numero
        }

        $ws.Cells.Item($targetRow, 2).Value2 = [string]$item.generalita
        $ws.Cells.Item($targetRow, 3).Value2 = [string]$item.rilascio
        $ws.Cells.Item($targetRow, 4).Value2 = [string]$item.scadenza
        $ws.Cells.Item($targetRow, 5).Value2 = [string]$item.note

        $numeroPattern = "^\s*" + [regex]::Escape($numero.Trim()) + "(\D|$)"
        for ($r = $usedStart; $r -le $usedEnd; $r++) {
            if ($r -eq $targetRow) { continue }
            $numText = [string]$ws.Cells.Item($r, 1).Text
            if ([string]::IsNullOrWhiteSpace($numText)) { continue }
            if ($numText -match $numeroPattern) {
                $genText = [string]$ws.Cells.Item($r, 2).Text
                $rilText = [string]$ws.Cells.Item($r, 3).Text
                $scaText = [string]$ws.Cells.Item($r, 4).Text
                $noteText = [string]$ws.Cells.Item($r, 5).Text
                if (
                    -not [string]::IsNullOrWhiteSpace($genText) -or
                    -not [string]::IsNullOrWhiteSpace($rilText) -or
                    -not [string]::IsNullOrWhiteSpace($scaText) -or
                    -not [string]::IsNullOrWhiteSpace($noteText)
                ) {
                    $ws.Range($ws.Cells.Item($r, 1), $ws.Cells.Item($r, 5)).ClearContents()
                }
            }
        }
    }
    $wb.Save()
}
finally {
    if ($wb -ne $null) { $wb.Close($true) }
    if ($excel -ne $null) { $excel.Quit() }
}
"""
    with tempfile.TemporaryDirectory(prefix="passapp_invalidi_") as tmp_dir:
        tmp_path = Path(tmp_dir)
        payload_path = tmp_path / "rows.json"
        script_path = tmp_path / "update.ps1"
        payload_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
        script_path.write_text(ps_script, encoding="utf-8")
        result = subprocess.run(
            [
                "powershell",
                "-NoProfile",
                "-ExecutionPolicy",
                "Bypass",
                "-File",
                str(script_path),
                "-WorkbookPath",
                str(workbook_path),
                "-PayloadPath",
                str(payload_path),
            ],
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=180,
        )
        if result.returncode != 0:
            details = (result.stderr or result.stdout or "Errore sconosciuto").strip()
            raise RuntimeError(details)

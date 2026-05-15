import datetime
import json
import tempfile
import unittest
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import patch

import openpyxl

from core.ospitalita_excel import write_pending_with_excel_com as write_ospitalita_pending_with_excel_com
from core.pass_invalidi_excel import write_pending_to_xlsx


def _date_value(value):
    return value.date() if hasattr(value, "date") else value


class TestExcelWriters(unittest.TestCase):
    def test_pass_invalidi_xlsx_writer_updates_slot_and_appends(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            workbook_path = Path(tmp_dir) / "registro_invalidi.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["N", "GENERALITA", "RILASCIO", "SCADENZA", "NOTE"])
            ws.append([1, "", None, None, None])
            ws.append([2, "ESISTENTE", "01/01/2025", "01/01/2026", "nota"])
            wb.save(workbook_path)
            wb.close()

            write_pending_to_xlsx(
                workbook_path,
                [
                    {
                        "numero": 1,
                        "cognome": "ROSSI",
                        "nome_proprio": "Mario",
                        "indirizzo": "Via Roma 1",
                        "rilascio": "10/03/2026",
                        "scadenza": "10/03/2031",
                        "note": "slot riusato",
                    },
                    {
                        "numero": 3,
                        "cognome": "BIANCHI",
                        "nome_proprio": "Luisa",
                        "indirizzo": "Via Milano 2",
                        "rilascio": "11/03/2026",
                        "scadenza": "11/03/2031",
                        "note": "nuova riga",
                    },
                ],
            )

            saved = openpyxl.load_workbook(workbook_path, data_only=True)
            ws_saved = saved.active
            self.assertEqual(ws_saved.cell(row=2, column=1).value, 1)
            self.assertEqual(ws_saved.cell(row=2, column=2).value, "ROSSI MARIO - VIA ROMA 1")
            self.assertEqual(_date_value(ws_saved.cell(row=2, column=3).value), datetime.date(2026, 3, 10))
            self.assertEqual(_date_value(ws_saved.cell(row=2, column=4).value), datetime.date(2031, 3, 10))
            self.assertEqual(ws_saved.cell(row=2, column=5).value, "slot riusato")
            self.assertEqual(ws_saved.cell(row=4, column=1).value, 3)
            self.assertEqual(ws_saved.cell(row=4, column=2).value, "BIANCHI LUISA - VIA MILANO 2")
            saved.close()

    def test_ospitalita_excel_com_writer_builds_expected_payload(self):
        captured_payload = None

        def fake_run(command, **_kwargs):
            nonlocal captured_payload
            payload_path = Path(command[command.index("-PayloadPath") + 1])
            captured_payload = json.loads(payload_path.read_text(encoding="utf-8"))
            return SimpleNamespace(returncode=0, stdout="", stderr="")

        pending_records = [
            {
                "progressivo": "7",
                "protocollo": "123/2026",
                "data": "28/04/2026",
                "denunciante_nome": "Mario Rossi",
                "denunciante_indirizzo": "Via Roma 1",
                "straniero_nome": "Ana Verdi",
                "straniero_indirizzo": "Via Milano 2",
                "motivo": "OSPITA",
                "indirizzo_ospitalita": "Via Ospitalita 3",
            }
        ]

        with patch("core.ospitalita_excel.subprocess.run", side_effect=fake_run):
            write_ospitalita_pending_with_excel_com(Path("registro_ospitalita.xls"), pending_records)

        self.assertEqual(captured_payload, pending_records)


if __name__ == "__main__":
    unittest.main()

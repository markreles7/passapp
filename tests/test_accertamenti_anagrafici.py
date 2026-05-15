import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from core.accertamenti_anagrafici import (
    AccertamentoAnagrafico,
    delete_accertamento,
    import_accertamenti_from_excel,
    load_accertamenti,
    next_accertamento_numero,
    save_accertamenti,
    save_accertamenti_to_excel,
    upsert_accertamento,
    validate_accertamento,
)


class TestAccertamentiAnagrafici(unittest.TestCase):
    def test_save_load_upsert_delete_and_next_number(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "accertamenti.json"
            records = [
                AccertamentoAnagrafico(1, "Mario Rossi", "Via Roma 1"),
                AccertamentoAnagrafico(4, "Anna Verdi", "Via Milano 2"),
            ]

            save_accertamenti(records, path)
            self.assertEqual(len(load_accertamenti(path)), 2)
            self.assertEqual(next_accertamento_numero(load_accertamenti(path)), 5)

            item = AccertamentoAnagrafico(4, "Anna Verdi", "Via Torino 3", positivo_data="15/05/2026", positivo_ora="09:30")
            upsert_accertamento(item, path)
            loaded = load_accertamenti(path)
            self.assertEqual(len(loaded), 2)
            self.assertEqual(loaded[1].indirizzo, "Via Torino 3")
            self.assertEqual(loaded[1].stato, "completato")

            delete_accertamento(1, path)
            self.assertEqual([record.numero for record in load_accertamenti(path)], [4])

    def test_validation_requires_required_fields_and_complete_date_time_pairs(self):
        self.assertFalse(validate_accertamento(AccertamentoAnagrafico(1, "", "Via Roma 1"))[0])
        self.assertFalse(validate_accertamento(AccertamentoAnagrafico(1, "Mario Rossi", ""))[0])
        self.assertFalse(
            validate_accertamento(
                AccertamentoAnagrafico(1, "Mario Rossi", "Via Roma 1", primo_negativo_data="15/05/2026")
            )[0]
        )
        self.assertFalse(
            validate_accertamento(
                AccertamentoAnagrafico(1, "Mario Rossi", "Via Roma 1", positivo_data="2026/15/05", positivo_ora="09:30")
            )[0]
        )
        self.assertTrue(
            validate_accertamento(
                AccertamentoAnagrafico(1, "Mario Rossi", "Via Roma 1", positivo_data="15/05/2026", positivo_ora="09:30")
            )[0]
        )

    def test_malformed_json_returns_empty_and_creates_backup(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "accertamenti.json"
            path.write_text("{json non valido", encoding="utf-8")

            records = load_accertamenti(path)
            backups = list((Path(tmp_dir) / "backups" / "accertamenti_anagrafici").glob("accertamenti_malformed_*.json"))

        self.assertEqual(records, [])
        self.assertEqual(len(backups), 1)

    def test_import_from_excel_normalizes_attempts(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "accertamenti.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["N°", "NOMINATIVO", "INDIRIZZO", "NOTE", "1° Negativo", "", "2° Negativo", "", "3° Negativo", ""])
            sheet.append(["", "", "", "", "Data", "Ora", "Data", "Ora", "Data", "Ora"])
            sheet.append([1, "Mario Rossi", "Via Roma 1", "Sabato mattina", "2026-05-15", "17;33", "", "", "", ""])
            workbook.save(path)

            records = import_accertamenti_from_excel(path)

        self.assertEqual(len(records), 1)
        self.assertEqual(records[0].numero, 1)
        self.assertEqual(records[0].primo_negativo_data, "15/05/2026")
        self.assertEqual(records[0].primo_negativo_ora, "17:33")
        self.assertEqual(records[0].stato, "in corso")

    def test_save_to_excel_writes_positive_columns(self):
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "accertamenti.xlsx"
            records = [
                AccertamentoAnagrafico(
                    1,
                    "Mario Rossi",
                    "Via Roma 1",
                    positivo_data="15/05/2026",
                    positivo_ora="09:30",
                )
            ]

            save_accertamenti_to_excel(path, records)

            workbook = load_workbook(path)
            sheet = workbook.active
            self.assertEqual(sheet.cell(row=1, column=11).value, "POSITIVO")
            self.assertEqual(sheet.cell(row=2, column=11).value, "Data")
            self.assertEqual(sheet.cell(row=3, column=11).value, "15/05/2026")
            self.assertEqual(sheet.cell(row=3, column=12).value, "09:30")
            self.assertTrue(sheet.cell(row=3, column=1).font.strike)


if __name__ == "__main__":
    unittest.main()

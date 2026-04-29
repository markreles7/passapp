"""
PassApp - Modulo Pass Invalidi.
Legge i registri configurati in data/config.json.
"""

import tkinter as tk
from tkinter import ttk, messagebox
import glob
import json
import os
import re
import shutil
import threading
import datetime
import queue
import subprocess
import tempfile
from pathlib import Path

from app_config import load_config, resolve_path
from core.audit import log_audit_event
from core.backups import create_excel_backup
from core.dates import format_date, giorni_rimanenti, parse_date
from core.file_state import FileSnapshot, file_matches_snapshot
from core.logging_utils import setup_module_logger
from core.text_utils import normalize_for_match
from core.workcopies import create_working_copy

try:
    import openpyxl
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

APP_CONFIG = load_config()
PATHS = APP_CONFIG["paths"]
UI_CONFIG = APP_CONFIG["ui"]
THEME = UI_CONFIG["theme"]
PASS_INVALIDI_UI = UI_CONFIG["modules"]["pass_invalidi"]

LOG_FILE = resolve_path(PATHS["log_file"])
logger = setup_module_logger(__name__, LOG_FILE)

# --- CONFIGURAZIONE ---
CARTELLA_RETE = PATHS["pass_invalidi_network_folder"]
PATTERN_FILE = PATHS["pass_invalidi_pattern"]
GIORNI_SCADENZA = int(APP_CONFIG["pass_invalidi"]["warning_days"])
LOGO_PATH = resolve_path(PATHS["logo_path"])
WORK_COPY_DIR = resolve_path("data/workcopies/pass_invalidi")
AUTH_DOCS_DIR = PATHS.get("pass_invalidi_docs_folder") or os.path.join(CARTELLA_RETE, "Tesserini rilasciati")
AUTH_TEMPLATE_PATH = PATHS.get("pass_invalidi_authorization_template") or os.path.join(
    AUTH_DOCS_DIR,
    "Tesserino Disabili Guida.doc",
)
AUTH_CITY_DEFAULT = PATHS.get("pass_invalidi_authorization_city", "PEGOGNAGA")

# Colori tema bianco-latte
BG = THEME["bg"]
BG2 = THEME["bg2"]
SURFACE = THEME["surface"]
BORDER = THEME["border"]
ACCENT = PASS_INVALIDI_UI["accent"]
ACCENT_DARK = PASS_INVALIDI_UI["accent_dark"]
SUCCESS = THEME["success"]
WARNING = THEME["warning"]
DANGER = THEME["danger"]
TEXT = THEME["text"]
TEXT_MUTED = THEME["text_muted"]
TEXT_DIM = THEME["text_dim"]

# --- PARSING ---

def get_status(val):
    """Ritorna 'expired', 'soon', 'valid' o None."""
    d = parse_date(val)
    if d is None:
        return None
    oggi = datetime.date.today()
    diff = (d - oggi).days
    if diff < 0:
        return "expired"
    if diff <= GIORNI_SCADENZA:
        return "soon"
    return "valid"


def carica_file(path):
    """Legge un file Excel registro e ritorna lista di dict."""
    records = []
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception as e:
        logger.exception("Errore lettura file pass invalidi: %s", path)
        return records, str(e)

    nome_file = os.path.basename(path)

    for row in rows:
        if not row or row[0] is None:
            continue
        # Col0 = numero progressivo (deve essere un numero)
        try:
            num = int(float(str(row[0]).strip()))
        except (ValueError, TypeError):
            continue

        generalita = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        if not generalita or generalita.lower() in ("generalita", "generalita'", "nan", "none"):
            continue

        nome, indirizzo = generalita, ""
        rilascio = row[2] if len(row) > 2 else None
        scadenza = row[3] if len(row) > 3 else None
        note_start = 4

        maybe_indirizzo = row[2] if len(row) > 2 else None
        has_indirizzo_col = (
            isinstance(maybe_indirizzo, str)
            and maybe_indirizzo.strip() != ""
            and parse_date(maybe_indirizzo) is None
        )

        if has_indirizzo_col:
            indirizzo = str(maybe_indirizzo).strip()
            rilascio = row[3] if len(row) > 3 else None
            scadenza = row[4] if len(row) > 4 else None
            note_start = 6 if len(row) > 6 else 5
        elif " - " in generalita:
            parts = generalita.split(" - ", 1)
            nome, indirizzo = parts[0].strip(), parts[1].strip()
        else:
            m = re.search(
                r"\b(VIA|STR\.|STRADA|PIAZZA|P\.ZA|CORSO|VIALE|FRAZ\.|FRAZIONE|VICOLO|BORGO|CONTRADA|LOC\.)\b",
                generalita, re.IGNORECASE
            )
            if m:
                idx = m.start()
                nome = generalita[:idx].strip(" -")
                indirizzo = generalita[idx:].strip()

        note_raw = " | ".join(
            str(c).strip() for c in (row[note_start:] if len(row) > note_start else [])
            if c and str(c).strip() not in ("None", "nan", "")
        )

        name_tokens = [t for t in re.split(r"\s+", nome.strip()) if t]
        cognome = name_tokens[0].upper() if name_tokens else nome.upper()
        nome_proprio = " ".join(name_tokens[1:]).strip() if len(name_tokens) > 1 else ""

        records.append({
            "numero":    num,
            "nome":      nome.upper(),
            "cognome":   cognome,
            "nome_proprio": nome_proprio,
            "nato_il":   "",
            "indirizzo": indirizzo,
            "rilascio":  rilascio,
            "scadenza":  scadenza,
            "note":      note_raw,
            "source":    nome_file,
        })

    return records, None


def _extract_year_from_name(name: str) -> int:
    years = [int(m.group(0)) for m in re.finditer(r"(?:19|20)\d{2}", name)]
    return max(years) if years else -1


def _file_sort_key(path: str):
    base = os.path.basename(path)
    year = _extract_year_from_name(base)
    try:
        mtime = os.path.getmtime(path)
    except OSError:
        mtime = 0.0
    return (year, mtime, base.lower())


def carica_tutti():
    """Cerca tutti i file nella cartella di rete e li carica."""
    pattern = os.path.join(CARTELLA_RETE, PATTERN_FILE)
    files = sorted(glob.glob(pattern), key=_file_sort_key, reverse=True)
    all_records = []
    errori = []
    for f in files:
        recs, err = carica_file(f)
        if err:
            errori.append(f"{os.path.basename(f)}: {err}")
        else:
            all_records.extend(recs)
    return all_records, files, errori


# --- INTERFACCIA GRAFICA ---

class PassInvalidiFrame(tk.Frame):
    def __init__(self, parent, controller=None, show_back_button=True):
        super().__init__(parent, bg=BG)
        self.controller = controller
        self.show_back_button = bool(show_back_button and controller is not None)

        self.all_records = []
        self.filtered    = []
        self.filter_mode = tk.StringVar(value="tutti")
        self.sort_col    = None
        self.sort_asc    = True
        self._dep_warned = False
        self._load_seq   = 0
        self._all_records_loaded = []
        self._files_loaded: list[str] = []
        self._source_files: list[str] = []
        self._primary_source_file: str | None = None
        self._working_copy_file: Path | None = None
        self._source_file_snapshot: FileSnapshot | None = None
        self._pending_new_records: list[dict] = []
        self.year_var = tk.StringVar(value="Tutti")
        self._last_year_value = "Tutti"

        self._build_ui()
        self.after(200, self.carica_dati)

    def on_show(self):
        if self.controller is not None:
            self.controller.title(PASS_INVALIDI_UI["title"])

    # --- UI ---

    def _build_ui(self):
        # HEADER
        hdr = tk.Frame(self, bg=SURFACE, relief="flat")
        hdr.pack(fill="x", side="top")

        tk.Frame(hdr, bg=BORDER, height=1).pack(fill="x", side="bottom")

        inner_hdr = tk.Frame(hdr, bg=SURFACE)
        inner_hdr.pack(fill="x", padx=24, pady=14)

        title_frame = tk.Frame(inner_hdr, bg=SURFACE)
        title_frame.pack(side="left")

        # Badge icona
        badge = tk.Label(title_frame, text="PI", bg=ACCENT, fg="white",
                         font=("Segoe UI", 16), width=2, padx=6, pady=4)
        badge.pack(side="left", padx=(0, 12))

        lbl_title = tk.Label(title_frame, text="Gestione Pass Invalidi",
                             bg=SURFACE, fg=TEXT,
                             font=("Segoe UI", 14, "bold"))
        lbl_title.pack(side="left")

        lbl_sub = tk.Label(title_frame, text=" - Polizia Locale",
                           bg=SURFACE, fg=TEXT_MUTED,
                           font=("Segoe UI", 10))
        lbl_sub.pack(side="left", pady=(3, 0))

        # Azioni header (destra)
        actions_frame = tk.Frame(inner_hdr, bg=SURFACE)
        actions_frame.pack(side="right")
        if self.show_back_button:
            tk.Button(actions_frame, text="<- Torna al Menu Principale",
                      bg=BG2, fg=TEXT_MUTED,
                      font=("Segoe UI", 10, "bold"),
                      relief="flat", cursor="hand2",
                      activebackground=BORDER,
                      highlightbackground=BORDER,
                      highlightthickness=1,
                      padx=12, pady=8,
                      command=self.torna_menu).pack(side="right", padx=(8, 0))

        # Stats a destra
        self.stats_frame = tk.Frame(inner_hdr, bg=SURFACE)
        self.stats_frame.pack(side="right")

        self.lbl_totale   = self._stat_pill(self.stats_frame, "Totale: -")
        self.lbl_scaduti  = self._stat_pill(self.stats_frame, "Scaduti: -", DANGER)
        self.lbl_scadenza = self._stat_pill(self.stats_frame, "In scad.: -", WARNING)
        self.lbl_file     = self._stat_pill(self.stats_frame, "File: -")

        # BARRA RICERCA + FILTRI
        search_panel = tk.Frame(self, bg=BG)
        search_panel.pack(fill="x", padx=24, pady=(18, 0))

        # Riga ricerca
        search_row = tk.Frame(search_panel, bg=BG)
        search_row.pack(fill="x")

        search_container = tk.Frame(search_row, bg=SURFACE,
                                    highlightbackground=BORDER,
                                    highlightthickness=1,
                                    relief="flat")
        search_container.pack(side="left", fill="x", expand=True)

        tk.Label(search_container, text="Cerca", bg=SURFACE,
                 font=("Segoe UI", 13), padx=10).pack(side="left")

        self.search_var = tk.StringVar()
        self.search_var.trace_add("write", lambda *_: self.applica_filtro())

        self.entry = tk.Entry(search_container, textvariable=self.search_var,
                              font=("Segoe UI", 13), relief="flat",
                              bg=SURFACE, fg=TEXT,
                              insertbackground=ACCENT,
                              bd=0)
        self.entry.pack(side="left", fill="x", expand=True, ipady=10)
        self.entry.bind("<Escape>", lambda e: self.search_var.set(""))

        tk.Button(search_container, text="x", bg=SURFACE, fg=TEXT_MUTED,
                  font=("Segoe UI", 10), relief="flat", cursor="hand2",
                  activebackground=BG2,
                  command=lambda: self.search_var.set("")).pack(side="right", padx=8)

        # Pulsante ricarica
        self.btn_reload = tk.Button(search_row, text="Aggiorna",
                                    bg=ACCENT, fg="white",
                                    font=("Segoe UI", 10, "bold"),
                                    relief="flat", cursor="hand2",
                                    activebackground=ACCENT_DARK,
                                    padx=16, pady=10,
                                    command=self.carica_dati)
        self.btn_reload.pack(side="left", padx=(10, 0))

        self.btn_new = tk.Button(
            search_row,
            text="+ Nuovo nominativo",
            bg=SUCCESS,
            fg="white",
            font=("Segoe UI", 10, "bold"),
            relief="flat",
            cursor="hand2",
            activebackground="#197A52",
            padx=14,
            pady=10,
            command=self.nuovo_nominativo,
            state="disabled",
        )
        self.btn_new.pack(side="left", padx=(8, 0))

        self.btn_edit = tk.Button(
            search_row,
            text="Modifica selezionato",
            bg=SUCCESS,
            fg="white",
            font=("Segoe UI", 10, "bold"),
            relief="flat",
            cursor="hand2",
            activebackground="#197A52",
            padx=14,
            pady=10,
            command=self.modifica_selezionato,
            state="disabled",
        )
        self.btn_edit.pack(side="left", padx=(8, 0))

        self.btn_save_changes = tk.Button(
            search_row,
            text="SALVA MODIFICHE",
            bg=WARNING,
            fg="white",
            font=("Segoe UI", 10, "bold"),
            relief="flat",
            cursor="hand2",
            activebackground="#9A5E0C",
            padx=14,
            pady=10,
            command=self.salva_modifiche,
            state="disabled",
        )
        self.btn_save_changes.pack(side="left", padx=(8, 0))

        year_wrap = tk.Frame(search_row, bg=BG)
        year_wrap.pack(side="left", padx=(10, 0))
        tk.Label(
            year_wrap,
            text="Anno",
            bg=BG,
            fg=TEXT_MUTED,
            font=("Segoe UI", 9, "bold"),
        ).pack(side="left", padx=(0, 6))
        self.year_combo = ttk.Combobox(
            year_wrap,
            textvariable=self.year_var,
            values=("Tutti",),
            state="readonly",
            width=10,
            font=("Segoe UI", 10),
        )
        self.year_combo.pack(side="left")
        self.year_combo.bind("<<ComboboxSelected>>", self._on_year_change)

        # Pulsante apri cartella
        tk.Button(search_row, text="Apri cartella",
                  bg=BG2, fg=TEXT_MUTED,
                  font=("Segoe UI", 10),
                  relief="flat", cursor="hand2",
                  activebackground=BORDER,
                  highlightbackground=BORDER,
                  highlightthickness=1,
                  padx=14, pady=10,
                  command=self.apri_cartella).pack(side="left", padx=(8, 0))

        # Riga filtri
        filter_row = tk.Frame(search_panel, bg=BG)
        filter_row.pack(fill="x", pady=(10, 0))

        tk.Label(filter_row, text="Filtro:", bg=BG, fg=TEXT_MUTED,
                 font=("Segoe UI", 9)).pack(side="left", padx=(2, 8))

        self.filter_btns = {}
        filtri = [
            ("tutti",   "Tutti",                None),
            ("valid",   "Validi",             SUCCESS),
            ("soon",    "In scadenza (60gg)", WARNING),
            ("expired", "Scaduti",            DANGER),
        ]
        for key, label, color in filtri:
            btn = tk.Button(filter_row, text=label,
                            font=("Segoe UI", 9),
                            relief="flat", cursor="hand2",
                            padx=12, pady=5,
                            command=lambda k=key: self.set_filter(k))
            btn.pack(side="left", padx=3)
            self.filter_btns[key] = btn

        self._aggiorna_filter_btns()

        # STATUS BAR
        self.lbl_status = tk.Label(search_panel, text="Caricamento in corso...",
                                   bg=BG, fg=TEXT_DIM, font=("Segoe UI", 8),
                                   anchor="w")
        self.lbl_status.pack(fill="x", pady=(8, 0))

        # TABELLA
        table_frame = tk.Frame(self, bg=BG)
        table_frame.pack(fill="both", expand=True, padx=24, pady=(12, 24))

        # Scrollbar verticale
        vsb = ttk.Scrollbar(table_frame, orient="vertical")
        vsb.pack(side="right", fill="y")

        # Stile Treeview
        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure("Treeview",
                        background=SURFACE,
                        fieldbackground=SURFACE,
                        foreground=TEXT,
                        rowheight=34,
                        font=("Segoe UI", 10),
                        bordercolor=BORDER,
                        relief="flat")
        style.configure("Treeview.Heading",
                        background=BG2,
                        foreground=TEXT_MUTED,
                        font=("Segoe UI", 9, "bold"),
                        relief="flat",
                        bordercolor=BORDER)
        style.map("Treeview",
                  background=[("selected", "#DBEAFE")],
                  foreground=[("selected", ACCENT_DARK)])
        style.map("Treeview.Heading",
                  background=[("active", BORDER)])

        cols = ("numero", "nome", "indirizzo", "rilascio", "scadenza", "stato", "note", "source")
        self.tree = ttk.Treeview(table_frame, columns=cols, show="headings",
                                  yscrollcommand=vsb.set, selectmode="browse")
        vsb.config(command=self.tree.yview)

        # Intestazioni
        hdrs = {
            "numero":    ("N°",             52,  "center"),
            "nome":      ("Cognome e Nome", 230, "w"),
            "indirizzo": ("Indirizzo",      220, "w"),
            "rilascio":  ("Rilascio",       90,  "center"),
            "scadenza":  ("Scadenza",       90,  "center"),
            "stato":     ("Stato",          110, "center"),
            "note":      ("Note / Certificato", 200, "w"),
            "source":    ("File",           180, "w"),
        }
        for col, (lbl, w, anchor) in hdrs.items():
            self.tree.heading(col, text=lbl,
                              command=lambda c=col: self.ordina(c))
            self.tree.column(col, width=w, anchor=anchor, minwidth=40)

        # Tag colori riga
        self.tree.tag_configure("expired", background="#FDF0EF", foreground=DANGER)
        self.tree.tag_configure("soon",    background="#FFFBF4", foreground=WARNING)
        self.tree.tag_configure("valid",   background=SURFACE,  foreground=TEXT)
        self.tree.tag_configure("unknown", background=SURFACE,  foreground=TEXT_MUTED)
        self.tree.tag_configure("odd",     background="#FAFAF8")

        self.tree.pack(fill="both", expand=True)
        self.tree.bind("<Double-1>", self.mostra_dettaglio)

        # Hint doppio click
        tk.Label(self, text="Doppio clic su una riga per vedere tutti i dettagli",
                 bg=BG, fg=TEXT_DIM, font=("Segoe UI", 8)).pack(pady=(0, 6))

        # BANNER LOGO IN FONDO AL CENTRO
        self.banner_frame = tk.Frame(self, bg=BG)
        self.banner_frame.pack(fill="x", side="bottom")
        self._build_logo_banner()

    def _build_logo_banner(self):
        """Banner logo centrato in fondo, semi-trasparente."""
        try:
            from PIL import Image, ImageTk

            img = Image.open(LOGO_PATH).convert("RGBA")

            h_target = 100
            ratio = h_target / img.height
            new_w = int(img.width * ratio)
            img = img.resize((new_w, h_target), Image.LANCZOS)

            r, g, b, a = img.split()
            a = a.point(lambda x: int(x * 0.45))
            img.putalpha(a)

            bg_img = Image.new("RGBA", img.size, (245, 243, 239, 255))
            bg_img.paste(img, mask=img.split()[3])

            self._banner_img = ImageTk.PhotoImage(bg_img.convert("RGB"))

            tk.Frame(self.banner_frame, bg=BORDER, height=1).pack(fill="x")

            lbl = tk.Label(self.banner_frame, image=self._banner_img, bg=BG, bd=0)
            lbl.pack(pady=(4, 6))
        except Exception:
            pass  # Pillow non disponibile o logo assente

    def _stat_pill(self, parent, text, color=None):
        f = tk.Frame(parent, bg=BG2, relief="flat",
                     highlightbackground=BORDER, highlightthickness=1)
        f.pack(side="left", padx=4)
        lbl = tk.Label(f, text=text, bg=BG2,
                       fg=color or TEXT_MUTED,
                       font=("Segoe UI", 9, "bold"), padx=10, pady=4)
        lbl.pack()
        return lbl

    @staticmethod
    def _year_from_label(text: str) -> str:
        year = _extract_year_from_name(text)
        return str(year) if year > 0 else ""

    def _file_year(self, path: str) -> str:
        return self._year_from_label(os.path.basename(path))

    def _record_year(self, rec: dict) -> str:
        return self._year_from_label(str(rec.get("source", "")))

    def _refresh_year_choices(self):
        years = sorted(
            {self._file_year(path) for path in self._files_loaded if self._file_year(path)},
            reverse=True,
        )
        values = ["Tutti"] + years if years else ["Tutti"]
        self.year_combo.configure(values=values)

        current = self.year_var.get().strip()
        if current in values and current != "Tutti":
            selected = current
        else:
            selected = years[0] if years else "Tutti"
        self.year_var.set(selected)
        self._last_year_value = selected

    def _refresh_summary_stats(self):
        records = list(self.all_records)
        n_scad = sum(1 for r in records if get_status(r.get("scadenza")) == "expired")
        n_soon = sum(1 for r in records if get_status(r.get("scadenza")) == "soon")
        self.lbl_totale.config(text=f"Totale: {len(records)}")
        self.lbl_file.config(text=f"File: {len(self._source_files)}")
        self.lbl_scaduti.config(text=f"Scaduti: {n_scad}")
        self.lbl_scadenza.config(text=f"In scad.: {n_soon}")

    def _apply_year_context(self):
        selected_year = self.year_var.get().strip()
        selected_files = list(self._files_loaded)
        selected_records = list(self._all_records_loaded)

        if selected_year and selected_year != "Tutti":
            selected_files = [path for path in self._files_loaded if self._file_year(path) == selected_year]
            source_names = {os.path.basename(path) for path in selected_files}
            selected_records = [rec for rec in self._all_records_loaded if str(rec.get("source", "")) in source_names]

        self._source_files = selected_files
        self.all_records = selected_records
        if selected_year == "Tutti" and len(selected_files) != 1:
            self._primary_source_file = None
        else:
            self._primary_source_file = selected_files[0] if selected_files else None

        self._prepare_working_copy()
        self.btn_new.config(state="normal" if self._working_copy_file else "disabled")
        self.btn_edit.config(state="normal" if self._working_copy_file else "disabled")
        self._refresh_summary_stats()

    def _on_year_change(self, _event=None):
        next_year = self.year_var.get().strip() or "Tutti"
        prev_year = self._last_year_value
        if next_year == prev_year:
            return

        if self._pending_new_records and not self._confirm_pending_before_context_switch():
            self.year_var.set(prev_year)
            return

        self._last_year_value = next_year
        self._pending_new_records = []
        self.btn_save_changes.config(state="disabled")
        self._apply_year_context()
        self.applica_filtro()

    # --- DATI ---

    def carica_dati(self, force: bool = False):
        if not force and self._pending_new_records:
            if not self._confirm_pending_before_context_switch():
                return

        if not OPENPYXL_OK:
            if not self._dep_warned:
                messagebox.showerror(
                    "Dipendenza mancante",
                    "La libreria 'openpyxl' non e installata.\nEsegui: pip install openpyxl",
                )
                self._dep_warned = True
            logger.error("Dipendenza openpyxl non disponibile durante caricamento pass invalidi")
            self.lbl_status.config(text="Errore: dipendenza openpyxl non disponibile.")
            return

        self.btn_reload.config(state="disabled", text="Caricamento...")
        self.lbl_status.config(text=f"Lettura da {CARTELLA_RETE} ...")
        self.tree.delete(*self.tree.get_children())
        self._load_seq += 1
        current_seq = self._load_seq
        result_q: queue.Queue = queue.Queue()

        def worker():
            try:
                records, files, errori = carica_tutti()
            except Exception as exc:
                logger.exception("Errore imprevisto nel worker di caricamento pass invalidi")
                records, files, errori = [], [], [str(exc)]
            result_q.put((records, files, errori))

        def poll():
            if current_seq != self._load_seq:
                return
            if not self.winfo_exists():
                return
            try:
                records, files, errori = result_q.get_nowait()
            except queue.Empty:
                try:
                    self.after(60, poll)
                except tk.TclError:
                    pass
                return

            if current_seq != self._load_seq or not self.winfo_exists():
                return
            self._post_carica_v2(records, files, errori)

        threading.Thread(target=worker, daemon=True).start()
        self.after(60, poll)

    def _post_carica_v2(self, records, files, errori):
        self.btn_reload.config(state="normal", text="Aggiorna")
        self._all_records_loaded = list(records)
        self._files_loaded = list(files)

        if not files and not records:
            logger.warning("Nessun file pass invalidi disponibile in %s", CARTELLA_RETE)
            self.all_records = []
            self._all_records_loaded = []
            self._files_loaded = []
            self.filtered = []
            self._source_files = []
            self._primary_source_file = None
            self._working_copy_file = None
            self._pending_new_records = []
            self._popola_tabella([])
            self.lbl_totale.config(text="Totale: 0")
            self.lbl_file.config(text="File: 0")
            self.lbl_scaduti.config(text="Scaduti: 0")
            self.lbl_scadenza.config(text="In scad.: 0")
            self.btn_new.config(state="disabled")
            self.btn_edit.config(state="disabled")
            self.btn_save_changes.config(state="disabled")
            self.year_combo.configure(values=("Tutti",))
            self.year_var.set("Tutti")
            self._last_year_value = "Tutti"
            msg = (
                f"Nessun file trovato in:\n{CARTELLA_RETE}\n\n"
                f"Verifica che il disco R: sia collegato e che esistano\n"
                f"file con nome: {PATTERN_FILE}"
            )
            messagebox.showwarning("Cartella non trovata", msg)
            self.lbl_status.config(text="Nessun file trovato: controlla disco R:")
            return

        if errori:
            logger.error("Errori lettura pass invalidi: %s", " | ".join(errori))
            messagebox.showerror(
                "Errori di lettura",
                "Alcuni file non sono stati letti correttamente:\n\n" + "\n".join(errori),
            )

        self.all_records = list(records)
        self._pending_new_records = []
        self._refresh_year_choices()
        self._apply_year_context()

        nomi_file = ", ".join(os.path.basename(f) for f in self._source_files)
        selected_year = self.year_var.get().strip()
        year_label = f"Anno {selected_year}" if selected_year and selected_year != "Tutti" else "Tutti gli anni"
        self.lbl_status.config(
            text=(
                f"{year_label} - Caricati {len(self._source_files)} file - "
                f"{len(self.all_records)} pass totali - Aggiornato: "
                f"{datetime.datetime.now().strftime('%H:%M:%S')} | {nomi_file}"
            )
        )
        self.btn_new.config(state="normal" if self._working_copy_file else "disabled")
        self.btn_edit.config(state="normal" if self._working_copy_file else "disabled")
        self.btn_save_changes.config(state="disabled")
        self.applica_filtro()

    def applica_filtro(self, *_):
        query = self.search_var.get().strip().lower()
        mode  = self.filter_mode.get()
        parts = query.split() if query else []

        risultati = []
        for r in self.all_records:
            nome_lower = r["nome"].lower()
            if parts and not all(p in nome_lower for p in parts):
                continue
            st = get_status(r["scadenza"])
            if mode == "expired" and st != "expired":
                continue
            if mode == "soon" and st != "soon":
                continue
            if mode == "valid" and st != "valid":
                continue
            risultati.append(r)

        self.filtered = risultati
        self._popola_tabella(risultati)

        n = len(risultati)
        self.lbl_status.config(
            text=(f"{'Tutti i' if not parts else 'Trovati'} {n} risultati"
                  + (f' per "{query}"' if parts else "")
                  + (f" - filtro: {mode}" if mode != "tutti" else ""))
        )

    def _popola_tabella(self, records):
        self.tree.delete(*self.tree.get_children())

        for i, r in enumerate(records):
            st = get_status(r["scadenza"])
            tag = st or "unknown"

            # Etichetta stato
            if st == "expired":
                giorni = giorni_rimanenti(r["scadenza"])
                stato_lbl = f"Scaduto ({abs(giorni)}gg fa)" if giorni is not None else "Scaduto"
            elif st == "soon":
                giorni = giorni_rimanenti(r["scadenza"])
                stato_lbl = f"Scade in {giorni}gg" if giorni is not None else "In scadenza"
            elif st == "valid":
                stato_lbl = "Valido"
            else:
                stato_lbl = "-"

            # Alterna sfondo righe per leggibilita (solo per valid/unknown)
            if tag in ("valid", "unknown") and i % 2 == 1:
                tag = "odd"

            self.tree.insert("", "end", iid=str(i), tags=(tag,), values=(
                r["numero"],
                r["nome"],
                r["indirizzo"] or "-",
                format_date(r["rilascio"]),
                format_date(r["scadenza"]),
                stato_lbl,
                r["note"] or "-",
                r["source"],
            ))

    # --- ORDINAMENTO ---

    def ordina(self, col):
        if self.sort_col == col:
            self.sort_asc = not self.sort_asc
        else:
            self.sort_col = col
            self.sort_asc = True

        def key(r):
            val = {
                "numero":    r["numero"],
                "nome":      r["nome"],
                "indirizzo": r["indirizzo"],
                "rilascio":  parse_date(r["rilascio"]) or datetime.date.min,
                "scadenza":  parse_date(r["scadenza"]) or datetime.date.min,
                "stato":     get_status(r["scadenza"]) or "",
                "note":      r["note"],
                "source":    r["source"],
            }.get(col, "")
            return val

        self.filtered.sort(key=key, reverse=not self.sort_asc)
        self._popola_tabella(self.filtered)

        # Aggiorna intestazione con freccia
        for c in ("numero", "nome", "indirizzo", "rilascio", "scadenza", "stato", "note", "source"):
            base = {"numero":"N°","nome":"Cognome e Nome","indirizzo":"Indirizzo",
                    "rilascio":"Rilascio","scadenza":"Scadenza","stato":"Stato",
                    "note":"Note / Certificato","source":"File"}[c]
            arrow = (" ^" if self.sort_asc else " v") if c == col else ""
            self.tree.heading(c, text=base + arrow)

    # --- FILTRI ---

    def apri_cartella(self):
        """Apre la cartella di rete in Esplora File."""
        import subprocess
        if os.path.isdir(CARTELLA_RETE):
            subprocess.Popen(f'explorer "{CARTELLA_RETE}"')
        else:
            messagebox.showwarning(
                "Cartella non trovata",
                f"Impossibile aprire:\n{CARTELLA_RETE}\n\nVerifica che il disco R: sia collegato."
            )

    def torna_menu(self):
        if self._pending_new_records:
            if not self._confirm_pending_before_context_switch():
                return
        if self.controller is not None and hasattr(self.controller, "show_frame"):
            self.controller.show_frame("MainMenuFrame")

    def set_filter(self, key):
        self.filter_mode.set(key)
        self._aggiorna_filter_btns()
        self.applica_filtro()

    def _aggiorna_filter_btns(self):
        colors = {"tutti": ACCENT, "valid": SUCCESS, "soon": WARNING, "expired": DANGER}
        cur = self.filter_mode.get()
        for key, btn in self.filter_btns.items():
            if key == cur:
                btn.config(bg=colors[key], fg="white",
                           activebackground=colors[key], activeforeground="white")
            else:
                btn.config(bg=SURFACE, fg=TEXT_MUTED,
                           activebackground=BG2, activeforeground=TEXT)

    def _prepare_working_copy(self):
        self._working_copy_file = None
        self._source_file_snapshot = None
        if not self._primary_source_file:
            return
        try:
            result = create_working_copy(self._primary_source_file, WORK_COPY_DIR, prefix="invalidi_")
            if result.removed_old_copies:
                logger.info("Rimosse %s copie di lavoro vecchie pass invalidi", result.removed_old_copies)
            self._working_copy_file = result.path
            self._source_file_snapshot = result.snapshot
        except OSError:
            logger.exception("Impossibile creare copia di lavoro pass invalidi")
            messagebox.showwarning(
                "Copia di lavoro non disponibile",
                "Non e stato possibile creare la copia di lavoro del registro.",
            )

    def _confirm_pending_before_context_switch(self) -> bool:
        answer = messagebox.askyesnocancel(
            "Modifiche non salvate",
            "Sono presenti modifiche non salvate.\nVuoi salvarle prima di continuare?",
            parent=self,
        )
        if answer is None:
            return False
        if answer:
            return self.salva_modifiche(trigger_reload=False)
        self._discard_pending_changes()
        return True

    def _next_numero(self) -> int:
        max_num = 0
        for rec in self.all_records:
            try:
                max_num = max(max_num, int(rec.get("numero", 0)))
            except (TypeError, ValueError):
                continue
        return max_num + 1

    def _append_pending_to_xlsx(self, workbook_path: Path):
        wb = openpyxl.load_workbook(workbook_path)
        ws = wb.active
        next_row = ws.max_row + 1

        def maybe_date(text):
            parsed = parse_date(text)
            return parsed if parsed is not None else text

        for rec in self._pending_new_records:
            ws.cell(row=next_row, column=1, value=rec["numero"])
            ws.cell(row=next_row, column=2, value=rec["nome"])
            ws.cell(row=next_row, column=3, value=rec["indirizzo"])
            ws.cell(row=next_row, column=4, value=maybe_date(rec["rilascio"]))
            ws.cell(row=next_row, column=5, value=maybe_date(rec["scadenza"]))
            ws.cell(row=next_row, column=7, value=rec["note"])
            next_row += 1
        wb.save(workbook_path)
        wb.close()

    def _append_pending_with_excel_com(self, workbook_path: Path):
        payload = [
            {
                "numero": rec.get("numero", ""),
                "nome": rec.get("nome", ""),
                "indirizzo": rec.get("indirizzo", ""),
                "rilascio": rec.get("rilascio", ""),
                "scadenza": rec.get("scadenza", ""),
                "note": rec.get("note", ""),
            }
            for rec in self._pending_new_records
        ]

        ps_script = r"""
param(
    [Parameter(Mandatory = $true)][string]$WorkbookPath,
    [Parameter(Mandatory = $true)][string]$PayloadPath
)
$ErrorActionPreference = "Stop"
$items = Get-Content -Path $PayloadPath -Raw -Encoding UTF8 | ConvertFrom-Json
$excel = $null
$wb = $null
$ws = $null
try {
    $excel = New-Object -ComObject Excel.Application
    $excel.Visible = $false
    $excel.DisplayAlerts = $false
    $wb = $excel.Workbooks.Open($WorkbookPath)
    $ws = $wb.Worksheets.Item(1)

    $last = $ws.Cells($ws.Rows.Count, 1).End(-4162).Row
    if ($last -lt 1) { $last = 1 }
    if (-not $ws.Cells.Item($last, 1).Value2) { $last = $last - 1 }
    $row = $last + 1

    foreach ($item in $items) {
        $ws.Cells.Item($row, 1).Value2 = [string]$item.numero
        $ws.Cells.Item($row, 2).Value2 = [string]$item.nome
        $ws.Cells.Item($row, 3).Value2 = [string]$item.indirizzo
        $ws.Cells.Item($row, 4).Value2 = [string]$item.rilascio
        $ws.Cells.Item($row, 5).Value2 = [string]$item.scadenza
        $ws.Cells.Item($row, 7).Value2 = [string]$item.note
        $row += 1
    }

    $wb.Save()
}
finally {
    if ($wb -ne $null) { $wb.Close($true) }
    if ($excel -ne $null) { $excel.Quit() }
}
"""
        with tempfile.TemporaryDirectory(prefix="passapp_invalidi_") as tmp_dir:
            tmp_path = Path(tmp_dir)
            payload_path = tmp_path / "rows.json"
            script_path = tmp_path / "append.ps1"
            payload_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
            script_path.write_text(ps_script, encoding="utf-8")
            result = subprocess.run(
                [
                    "powershell",
                    "-NoProfile",
                    "-ExecutionPolicy",
                    "Bypass",
                    "-File",
                    str(script_path),
                    "-WorkbookPath",
                    str(workbook_path),
                    "-PayloadPath",
                    str(payload_path),
                ],
                capture_output=True,
                text=True,
                encoding="utf-8",
                errors="replace",
                timeout=180,
            )
            if result.returncode != 0:
                details = (result.stderr or result.stdout or "Errore sconosciuto").strip()
                raise RuntimeError(details)

    @staticmethod
    def _normalize_text_for_match(value: str) -> str:
        return normalize_for_match(value)

    def _current_source_name(self) -> str:
        if not self._primary_source_file:
            return ""
        return os.path.basename(self._primary_source_file).strip()

    @staticmethod
    def _record_numero_int(record: dict) -> int | None:
        try:
            return int(str(record.get("numero", "")).strip())
        except (TypeError, ValueError):
            return None

    @staticmethod
    def _blankish(value) -> bool:
        text = str(value or "").strip().lower()
        return text in {"", "-", "none", "nan"}

    def _is_empty_slot(self, record: dict) -> bool:
        fields = (
            record.get("nome", ""),
            record.get("indirizzo", ""),
            record.get("rilascio", ""),
            record.get("scadenza", ""),
            record.get("note", ""),
        )
        return all(self._blankish(v) for v in fields)

    def _find_record_by_numero(self, numero: int, source_name: str | None = None) -> dict | None:
        source_ref = (source_name or self._current_source_name()).strip()
        for rec in self.all_records:
            if source_ref and str(rec.get("source", "")).strip() != source_ref:
                continue
            if self._record_numero_int(rec) == numero:
                return rec
        return None

    def _reserve_numero_slot(self) -> tuple[int, dict | None, dict | None, str]:
        source_ref = self._current_source_name()
        numeric_records: list[tuple[int, dict]] = []
        for rec in self.all_records:
            if source_ref and str(rec.get("source", "")).strip() != source_ref:
                continue
            num = self._record_numero_int(rec)
            if num is not None:
                numeric_records.append((num, rec))

        for num, rec in sorted(numeric_records, key=lambda item: item[0]):
            if rec.get("_pending"):
                continue
            if self._is_empty_slot(rec):
                return num, rec, dict(rec), "update_slot"

        max_num = max((num for num, _ in numeric_records), default=0)
        return max_num + 1, None, None, "append"

    def _selected_record(self) -> dict | None:
        sel = self.tree.selection()
        if not sel:
            return None
        try:
            idx = int(sel[0])
        except (TypeError, ValueError):
            return None
        if idx < 0 or idx >= len(self.filtered):
            return None
        return self.filtered[idx]

    def _upsert_pending_record(self, pending: dict):
        source = str(pending.get("source", "")).strip()
        numero = str(pending.get("numero", "")).strip()
        for idx, existing in enumerate(self._pending_new_records):
            if (
                str(existing.get("source", "")).strip() == source
                and str(existing.get("numero", "")).strip() == numero
            ):
                if existing.get("original_snapshot") and not pending.get("original_snapshot"):
                    pending["original_snapshot"] = existing.get("original_snapshot")
                self._pending_new_records[idx] = pending
                return
        self._pending_new_records.append(pending)

    def _discard_pending_changes(self):
        for pending in reversed(self._pending_new_records):
            mode = pending.get("mode")
            source = str(pending.get("source", "")).strip()
            try:
                numero = int(str(pending.get("numero", "")).strip())
            except (TypeError, ValueError):
                numero = None
            if numero is None:
                continue

            if mode in {"update_slot", "update_existing"}:
                target = self._find_record_by_numero(numero, source_name=source or None)
                original = pending.get("original_snapshot")
                if target is not None and isinstance(original, dict):
                    target.clear()
                    target.update(original)
            elif mode == "append":
                self.all_records = [
                    rec for rec in self.all_records
                    if not (
                        rec.get("_pending")
                        and self._record_numero_int(rec) == numero
                        and str(rec.get("source", "")).strip() == source
                    )
                ]

        self._pending_new_records = []
        self.btn_save_changes.config(state="disabled")
        self.applica_filtro()

    def modifica_selezionato(self, _event=None):
        if not self._working_copy_file:
            messagebox.showwarning(
                "Modifica non disponibile",
                "Per modificare i record seleziona un anno specifico.",
            )
            return
        record = self._selected_record()
        if record is None:
            messagebox.showinfo("Nessuna selezione", "Seleziona un nominativo da modificare.")
            return
        self.nuovo_nominativo(record_to_edit=record)

    def nuovo_nominativo(self, record_to_edit: dict | None = None):
        if not self._working_copy_file:
            messagebox.showwarning("Operazione non disponibile", "Nessuna copia di lavoro disponibile.")
            return

        is_edit = record_to_edit is not None
        win = tk.Toplevel(self)
        win.title("Modifica nominativo invalidi" if is_edit else "Nuovo nominativo invalidi")
        win.configure(bg=BG)
        win.resizable(False, False)
        win.grab_set()

        edit_record = record_to_edit or {}
        cognome_default = str(edit_record.get("cognome", "")).strip()
        nome_default = str(edit_record.get("nome_proprio", "")).strip()
        if not cognome_default and not nome_default:
            cognome_default, nome_default = self._split_cognome_nome(edit_record.get("nome", ""))

        vars_map = {
            "cognome": tk.StringVar(value=cognome_default),
            "nome": tk.StringVar(value=nome_default),
            "nato_il": tk.StringVar(value=format_date(edit_record.get("nato_il", "")) if edit_record else ""),
            "indirizzo": tk.StringVar(value=str(edit_record.get("indirizzo", "")).strip()),
            "rilascio": tk.StringVar(value=format_date(edit_record.get("rilascio", "")) if edit_record else ""),
            "scadenza": tk.StringVar(value=format_date(edit_record.get("scadenza", "")) if edit_record else ""),
            "note": tk.StringVar(value=str(edit_record.get("note", "")).strip()),
        }

        body = tk.Frame(win, bg=SURFACE, highlightbackground=BORDER, highlightthickness=1)
        body.pack(fill="both", expand=True, padx=14, pady=14)

        fields = [
            ("Cognome", "cognome"),
            ("Nome", "nome"),
            ("Nato il (GG/MM/AAAA)", "nato_il"),
            ("Indirizzo di residenza", "indirizzo"),
            ("Rilascio (GG/MM/AAAA)", "rilascio"),
            ("Scadenza (GG/MM/AAAA)", "scadenza"),
            ("Note", "note"),
        ]
        for label, key in fields:
            row = tk.Frame(body, bg=SURFACE)
            row.pack(fill="x", padx=12, pady=6)
            tk.Label(
                row,
                text=label,
                bg=SURFACE,
                fg=TEXT_MUTED,
                font=("Segoe UI", 9, "bold"),
                width=28,
                anchor="w",
            ).pack(side="left")
            tk.Entry(
                row,
                textvariable=vars_map[key],
                font=("Segoe UI", 10),
                bg="white",
                fg=TEXT,
                relief="solid",
                bd=1,
            ).pack(side="left", fill="x", expand=True, ipady=4)

        hint = tk.Label(
            body,
            text="La scadenza viene proposta automaticamente da rilascio + data nascita. Puoi correggerla prima del salvataggio.",
            bg=SURFACE,
            fg=TEXT_DIM,
            font=("Segoe UI", 8),
            anchor="w",
            justify="left",
        )
        hint.pack(fill="x", padx=12, pady=(0, 6))

        def refresh_scadenza_auto(*_args):
            data_rilascio = parse_date(vars_map["rilascio"].get().strip())
            data_nascita = parse_date(vars_map["nato_il"].get().strip())
            if data_rilascio is None or data_nascita is None:
                return
            auto_scadenza = self._compute_authorization_expiry(data_rilascio, data_nascita)
            vars_map["scadenza"].set(auto_scadenza.strftime("%d/%m/%Y"))

        vars_map["rilascio"].trace_add("write", refresh_scadenza_auto)
        vars_map["nato_il"].trace_add("write", refresh_scadenza_auto)

        footer = tk.Frame(body, bg=SURFACE)
        footer.pack(fill="x", padx=12, pady=(8, 12))

        def conferma():
            cognome = vars_map["cognome"].get().strip().upper()
            nome_proprio = vars_map["nome"].get().strip()
            nato_il_text = vars_map["nato_il"].get().strip()
            indirizzo = vars_map["indirizzo"].get().strip()
            rilascio_text = vars_map["rilascio"].get().strip()
            scadenza_text = vars_map["scadenza"].get().strip()
            note = vars_map["note"].get().strip()

            if not cognome:
                messagebox.showwarning("Dati incompleti", "Inserisci il cognome.", parent=win)
                return
            if not nome_proprio:
                messagebox.showwarning("Dati incompleti", "Inserisci il nome.", parent=win)
                return
            if not indirizzo:
                messagebox.showwarning("Dati incompleti", "Inserisci l'indirizzo di residenza.", parent=win)
                return

            data_nascita = parse_date(nato_il_text)
            if data_nascita is None:
                messagebox.showwarning("Data non valida", "Data nascita non valida. Usa GG/MM/AAAA.", parent=win)
                return

            data_rilascio = parse_date(rilascio_text)
            if data_rilascio is None:
                messagebox.showwarning("Data non valida", "Data rilascio non valida. Usa GG/MM/AAAA.", parent=win)
                return

            data_scadenza = parse_date(scadenza_text)
            if data_scadenza is None:
                messagebox.showwarning("Data non valida", "Data scadenza non valida. Usa GG/MM/AAAA.", parent=win)
                return

            auto_scadenza = self._compute_authorization_expiry(data_rilascio, data_nascita)
            auto_text = auto_scadenza.strftime("%d/%m/%Y")
            current_text = data_scadenza.strftime("%d/%m/%Y")
            if current_text != auto_text:
                keep_custom = messagebox.askyesno(
                    "Conferma scadenza",
                    (
                        f"Scadenza calcolata automaticamente: {auto_text}\n"
                        f"Scadenza inserita: {current_text}\n\n"
                        "Vuoi mantenere la scadenza inserita?"
                    ),
                    parent=win,
                )
                if not keep_custom:
                    return
            else:
                if not messagebox.askyesno(
                    "Conferma scadenza",
                    f"Scadenza automatica proposta: {auto_text}\nConfermi questa scadenza?",
                    parent=win,
                ):
                    return

            if not messagebox.askyesno(
                "Conferma modifica" if is_edit else "Conferma inserimento",
                (
                    "Vuoi registrare la modifica nella copia di lavoro?"
                    if is_edit
                    else "Vuoi salvare questo inserimento nella copia di lavoro?"
                ),
                parent=win,
            ):
                return

            if is_edit and record_to_edit is not None:
                numero = self._record_numero_int(record_to_edit)
                if numero is None:
                    messagebox.showerror("Errore record", "Numero progressivo non valido.", parent=win)
                    return
                source_name = str(record_to_edit.get("source", "")).strip() or self._current_source_name()
                target_record = record_to_edit
                original_snapshot = dict(record_to_edit)
                mode = "update_existing"
            else:
                numero, target_record, original_snapshot, mode = self._reserve_numero_slot()
                source_name = self._current_source_name()

            full_name_upper = self._compose_full_name(cognome, nome_proprio)
            if target_record is None:
                target_record = {"numero": numero, "source": source_name, "_pending": True}
                self.all_records.append(target_record)
            else:
                target_record["_pending"] = True

            target_record.update(
                {
                    "numero": numero,
                    "cognome": cognome,
                    "nome_proprio": nome_proprio.strip(),
                    "nome": full_name_upper,
                    "nato_il": data_nascita.strftime("%d/%m/%Y"),
                    "indirizzo": indirizzo,
                    "rilascio": data_rilascio.strftime("%d/%m/%Y"),
                    "scadenza": data_scadenza.strftime("%d/%m/%Y"),
                    "note": note,
                    "source": source_name,
                }
            )

            self._upsert_pending_record(
                {
                    "mode": mode,
                    "source": source_name,
                    "numero": numero,
                    "cognome": cognome,
                    "nome_proprio": nome_proprio.strip(),
                    "nome": full_name_upper,
                    "nato_il": data_nascita.strftime("%d/%m/%Y"),
                    "indirizzo": indirizzo,
                    "rilascio": data_rilascio.strftime("%d/%m/%Y"),
                    "scadenza": data_scadenza.strftime("%d/%m/%Y"),
                    "note": note,
                    "original_snapshot": original_snapshot,
                }
            )
            log_audit_event(
                "pass_invalidi",
                "update" if is_edit else "create",
                "pass_invalidi",
                str(numero),
                "Modifica nominativo registrata nella copia di lavoro"
                if is_edit
                else "Nuovo nominativo registrato nella copia di lavoro",
                extra={"source": source_name},
            )
            self.btn_save_changes.config(state="normal")
            self.applica_filtro()
            win.destroy()
            messagebox.showinfo(
                "Modifica registrata" if is_edit else "Inserimento registrato",
                (
                    "Record aggiornato nella copia di lavoro.\nUsa 'SALVA MODIFICHE' per aggiornare il file Excel."
                    if is_edit
                    else "Nominativo aggiunto nella copia di lavoro.\nUsa 'SALVA MODIFICHE' per aggiornare il file Excel."
                ),
            )

        tk.Button(
            footer,
            text="Salva modifica" if is_edit else "Salva inserimento",
            bg=ACCENT,
            fg="white",
            font=("Segoe UI", 10, "bold"),
            relief="flat",
            cursor="hand2",
            activebackground=ACCENT_DARK,
            padx=14,
            pady=8,
            command=conferma,
        ).pack(side="left")
        tk.Button(
            footer,
            text="Annulla",
            bg=BG2,
            fg=TEXT_MUTED,
            font=("Segoe UI", 10, "bold"),
            relief="flat",
            cursor="hand2",
            activebackground=BORDER,
            highlightbackground=BORDER,
            highlightthickness=1,
            padx=14,
            pady=8,
            command=win.destroy,
        ).pack(side="left", padx=(8, 0))

    def _write_pending_to_xlsx(self, workbook_path: Path):
        wb = openpyxl.load_workbook(workbook_path)
        ws = wb.active

        wb_values = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
        ws_values = wb_values.active
        row_by_numero: dict[int, int] = {}
        empty_rows: list[int] = []
        next_append_row = ws.max_row + 1
        for row_idx in range(1, ws_values.max_row + 1):
            num_val = ws_values.cell(row=row_idx, column=1).value
            text_val = ws_values.cell(row=row_idx, column=2).value
            try:
                numero = int(float(str(num_val).strip()))
            except (TypeError, ValueError):
                numero = None
            if numero is not None:
                row_by_numero[numero] = row_idx
                if text_val is None or str(text_val).strip() == "":
                    empty_rows.append(row_idx)
        wb_values.close()

        def maybe_date(text):
            parsed = parse_date(text)
            return parsed if parsed is not None else text

        for rec in self._pending_new_records:
            try:
                numero = int(str(rec.get("numero", "")).strip())
            except (TypeError, ValueError):
                continue

            target_row = row_by_numero.get(numero)
            if target_row is None:
                if empty_rows:
                    target_row = empty_rows.pop(0)
                else:
                    target_row = next_append_row
                    next_append_row += 1
                ws.cell(row=target_row, column=1, value=numero)
                row_by_numero[numero] = target_row
            elif target_row in empty_rows:
                empty_rows.remove(target_row)

            generalita = self._compose_generalita_indirizzo(
                rec.get("cognome", ""),
                rec.get("nome_proprio", ""),
                rec.get("indirizzo", ""),
            )

            ws.cell(row=target_row, column=2, value=generalita)
            ws.cell(row=target_row, column=3, value=maybe_date(rec.get("rilascio", "")))
            ws.cell(row=target_row, column=4, value=maybe_date(rec.get("scadenza", "")))
            ws.cell(row=target_row, column=5, value=rec.get("note", ""))

        wb.save(workbook_path)
        wb.close()

    def _write_pending_with_excel_com(self, workbook_path: Path):
        wb_values = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
        ws_values = wb_values.active

        def numero_da_cella(value) -> int | None:
            if value is None:
                return None
            if isinstance(value, bool):
                return None
            if isinstance(value, int):
                return value
            if isinstance(value, float):
                try:
                    return int(value)
                except (TypeError, ValueError):
                    return None
            txt = str(value).strip()
            if not txt:
                return None
            m = re.match(r"^(\d+)(?:\s|$)", txt)
            if not m:
                return None
            try:
                return int(m.group(1))
            except (TypeError, ValueError):
                return None

        row_by_numero: dict[int, int] = {}
        empty_rows: list[int] = []
        last_numbered_row = 1
        for row_idx in range(1, ws_values.max_row + 1):
            numero = numero_da_cella(ws_values.cell(row=row_idx, column=1).value)
            if numero is None:
                continue
            if numero not in row_by_numero:
                row_by_numero[numero] = row_idx
            if row_idx > last_numbered_row:
                last_numbered_row = row_idx
            generalita = ws_values.cell(row=row_idx, column=2).value
            if generalita is None or str(generalita).strip() == "":
                empty_rows.append(row_idx)
        wb_values.close()

        payload = [
            {
                "numero": rec.get("numero", ""),
                "generalita": self._compose_generalita_indirizzo(
                    rec.get("cognome", ""),
                    rec.get("nome_proprio", ""),
                    rec.get("indirizzo", ""),
                ),
                "indirizzo": rec.get("indirizzo", ""),
                "rilascio": rec.get("rilascio", ""),
                "scadenza": rec.get("scadenza", ""),
                "note": rec.get("note", ""),
                "target_row": None,
                "set_numero": False,
            }
            for rec in self._pending_new_records
        ]

        for item in payload:
            try:
                numero = int(str(item.get("numero", "")).strip())
            except (TypeError, ValueError):
                continue

            target_row = row_by_numero.get(numero)
            set_numero = False
            if target_row is None:
                if empty_rows:
                    target_row = empty_rows.pop(0)
                else:
                    last_numbered_row += 1
                    target_row = last_numbered_row
                    set_numero = True
            elif target_row in empty_rows:
                empty_rows.remove(target_row)
            item["target_row"] = target_row
            item["set_numero"] = set_numero
            row_by_numero[numero] = target_row

        ps_script = r"""
param(
    [Parameter(Mandatory = $true)][string]$WorkbookPath,
    [Parameter(Mandatory = $true)][string]$PayloadPath
)
$ErrorActionPreference = "Stop"
$items = Get-Content -Path $PayloadPath -Raw -Encoding UTF8 | ConvertFrom-Json
$excel = $null
$wb = $null
$ws = $null
try {
    $excel = New-Object -ComObject Excel.Application
    $excel.Visible = $false
    $excel.DisplayAlerts = $false
    $wb = $excel.Workbooks.Open($WorkbookPath)
    $ws = $wb.Worksheets.Item(1)
    foreach ($item in $items) {
        $targetRow = 0
        [void][int]::TryParse([string]$item.target_row, [ref]$targetRow)
        $setNumero = $false
        try { $setNumero = [bool]$item.set_numero } catch {}
        $numero = [string]$item.numero
        $usedStart = $ws.UsedRange.Row
        $usedEnd = $ws.UsedRange.Row + $ws.UsedRange.Rows.Count - 1

        if ($targetRow -le 0) {
            for ($r = $usedStart; $r -le $usedEnd; $r++) {
                $value = [string]$ws.Cells.Item($r, 1).Text
                if ([string]::IsNullOrWhiteSpace($value)) { continue }
                if ($value.Trim() -eq $numero.Trim()) {
                    $targetRow = $r
                    break
                }
            }
        }
        if ($targetRow -le 0) {
            for ($r = $usedStart; $r -le $usedEnd; $r++) {
                $numText = [string]$ws.Cells.Item($r, 1).Text
                $genText = [string]$ws.Cells.Item($r, 2).Text
                if (-not [string]::IsNullOrWhiteSpace($numText) -and [string]::IsNullOrWhiteSpace($genText)) {
                    $targetRow = $r
                    break
                }
            }
        }
        if ($targetRow -le 0) {
            $targetRow = $usedEnd + 1
            $setNumero = $true
        }
        if ($setNumero -or [string]::IsNullOrWhiteSpace([string]$ws.Cells.Item($targetRow, 1).Text)) {
            $ws.Cells.Item($targetRow, 1).Value2 = [string]$item.numero
        }

        $ws.Cells.Item($targetRow, 2).Value2 = [string]$item.generalita
        $ws.Cells.Item($targetRow, 3).Value2 = [string]$item.rilascio
        $ws.Cells.Item($targetRow, 4).Value2 = [string]$item.scadenza
        $ws.Cells.Item($targetRow, 5).Value2 = [string]$item.note

        $numeroPattern = "^\s*" + [regex]::Escape($numero.Trim()) + "(\D|$)"
        for ($r = $usedStart; $r -le $usedEnd; $r++) {
            if ($r -eq $targetRow) { continue }
            $numText = [string]$ws.Cells.Item($r, 1).Text
            if ([string]::IsNullOrWhiteSpace($numText)) { continue }
            if ($numText -match $numeroPattern) {
                $genText = [string]$ws.Cells.Item($r, 2).Text
                $rilText = [string]$ws.Cells.Item($r, 3).Text
                $scaText = [string]$ws.Cells.Item($r, 4).Text
                $noteText = [string]$ws.Cells.Item($r, 5).Text
                if (
                    -not [string]::IsNullOrWhiteSpace($genText) -or
                    -not [string]::IsNullOrWhiteSpace($rilText) -or
                    -not [string]::IsNullOrWhiteSpace($scaText) -or
                    -not [string]::IsNullOrWhiteSpace($noteText)
                ) {
                    $ws.Range($ws.Cells.Item($r, 1), $ws.Cells.Item($r, 5)).ClearContents()
                }
            }
        }
    }
    $wb.Save()
}
finally {
    if ($wb -ne $null) { $wb.Close($true) }
    if ($excel -ne $null) { $excel.Quit() }
}
"""
        with tempfile.TemporaryDirectory(prefix="passapp_invalidi_") as tmp_dir:
            tmp_path = Path(tmp_dir)
            payload_path = tmp_path / "rows.json"
            script_path = tmp_path / "update.ps1"
            payload_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
            script_path.write_text(ps_script, encoding="utf-8")
            result = subprocess.run(
                [
                    "powershell",
                    "-NoProfile",
                    "-ExecutionPolicy",
                    "Bypass",
                    "-File",
                    str(script_path),
                    "-WorkbookPath",
                    str(workbook_path),
                    "-PayloadPath",
                    str(payload_path),
                ],
                capture_output=True,
                text=True,
                encoding="utf-8",
                errors="replace",
                timeout=180,
            )
            if result.returncode != 0:
                details = (result.stderr or result.stdout or "Errore sconosciuto").strip()
                raise RuntimeError(details)

    def salva_modifiche(self, trigger_reload: bool = True) -> bool:
        if not self._pending_new_records:
            messagebox.showinfo("Nessuna modifica", "Non ci sono modifiche da salvare.")
            return False
        if not self._working_copy_file or not self._primary_source_file:
            messagebox.showwarning("Salvataggio non disponibile", "Copia di lavoro non disponibile.")
            return False

        suffix = self._working_copy_file.suffix.lower()
        if suffix not in (".xlsx", ".xls"):
            messagebox.showwarning(
                "Formato non supportato",
                "Il salvataggio modifiche e supportato per file .xls/.xlsx nel modulo Invalidi.",
            )
            return False

        if not file_matches_snapshot(self._source_file_snapshot):
            messagebox.showwarning(
                "File originale modificato",
                "Il file originale è stato modificato dopo l’apertura della copia di lavoro. "
                "Ricaricare i dati prima di salvare.",
            )
            return False

        try:
            if suffix == ".xlsx":
                self._write_pending_to_xlsx(self._working_copy_file)
            else:
                self._write_pending_with_excel_com(self._working_copy_file)
            create_excel_backup(self._primary_source_file, "pass_invalidi")
            log_audit_event("pass_invalidi", "backup", "excel", None, "Backup file originale Pass Invalidi creato")
            shutil.copy2(self._working_copy_file, self._primary_source_file)
        except Exception as exc:
            logger.exception("Errore salvataggio modifiche pass invalidi")
            log_audit_event(
                "pass_invalidi",
                "save",
                "excel",
                None,
                "Salvataggio modifiche Pass Invalidi non riuscito",
                result="error",
                error=str(exc),
            )
            messagebox.showerror("Salvataggio non riuscito", f"Impossibile salvare le modifiche.\n\nDettagli:\n{exc}")
            return False

        saved_count = len(self._pending_new_records)
        for rec in self.all_records:
            rec.pop("_pending", None)
        self._pending_new_records = []
        self.btn_save_changes.config(state="disabled")
        if trigger_reload:
            self.carica_dati(force=True)
        else:
            self.applica_filtro()
        messagebox.showinfo("Salvataggio completato", "Le modifiche sono state salvate sul file Excel.")
        log_audit_event(
            "pass_invalidi",
            "save",
            "excel",
            None,
            "Salvate modifiche Pass Invalidi sul file originale",
            extra={"count": saved_count},
        )
        return True

    @staticmethod
    def _split_cognome_nome(full_name: str) -> tuple[str, str]:
        tokens = [tok for tok in re.split(r"\s+", str(full_name or "").strip()) if tok]
        if not tokens:
            return "", ""
        cognome = tokens[0].upper()
        nome = " ".join(tokens[1:]) if len(tokens) > 1 else ""
        nome = " ".join(part.capitalize() for part in nome.split()) if nome else ""
        return cognome, nome

    @staticmethod
    def _compose_full_name(cognome: str, nome: str) -> str:
        left = str(cognome or "").strip().upper()
        right = str(nome or "").strip().upper()
        if left and right:
            return f"{left} {right}"
        return left or right

    @staticmethod
    def _compose_generalita_indirizzo(cognome: str, nome: str, indirizzo: str) -> str:
        full_name = PassInvalidiFrame._compose_full_name(cognome, nome)
        address = str(indirizzo or "").strip().upper()
        if full_name and address:
            return f"{full_name} - {address}"
        return full_name or address

    @staticmethod
    def _compute_authorization_expiry(protocol_date: datetime.date, birth_date: datetime.date) -> datetime.date:
        plus_years = 5 if (protocol_date.month, protocol_date.day) <= (birth_date.month, birth_date.day) else 6
        target_year = protocol_date.year + plus_years
        day = birth_date.day
        month = birth_date.month
        while day > 0:
            try:
                return datetime.date(target_year, month, day)
            except ValueError:
                day -= 1
        return datetime.date(target_year, month, 1)

    def _resolve_authorization_template(self) -> Path | None:
        candidate = Path(AUTH_TEMPLATE_PATH)
        if candidate.exists():
            return candidate

        docs_dir = Path(AUTH_DOCS_DIR)
        if not docs_dir.exists():
            return None

        preferred_patterns = (
            "*Guida*.doc",
            "*guida*.doc",
            "*Template*.doc",
            "*template*.doc",
            "*Guida*.docx",
            "*guida*.docx",
            "*Template*.docx",
            "*template*.docx",
        )
        for pattern in preferred_patterns:
            matches = sorted(docs_dir.glob(pattern))
            if matches:
                return matches[0]

        candidates = sorted(list(docs_dir.glob("*.doc")) + list(docs_dir.glob("*.docx")))
        for path in candidates:
            normalized = self._normalize_text_for_match(path.stem)
            if "autoriz" not in normalized:
                continue
            if re.search(r"\bn\.?\s*\d+\b", normalized):
                continue
            return path
        return None

    def _authorization_year_for_record(self, record: dict) -> int:
        year_from_source = _extract_year_from_name(str(record.get("source", "")))
        if year_from_source > 0:
            return year_from_source
        date_rilascio = parse_date(record.get("rilascio"))
        if date_rilascio is not None:
            return date_rilascio.year
        return datetime.date.today().year

    def _find_authorization_doc_for_record(self, record: dict) -> Path | None:
        docs_dir = Path(AUTH_DOCS_DIR)
        if not docs_dir.exists():
            return None
        cognome = str(record.get("cognome", "")).strip()
        nome = str(record.get("nome_proprio", "")).strip()
        if not cognome and not nome:
            cognome, nome = self._split_cognome_nome(record.get("nome", ""))
        numero = str(record.get("numero", "")).strip()
        year = str(self._authorization_year_for_record(record))
        cognome_n = self._normalize_text_for_match(cognome)
        nome_n = self._normalize_text_for_match(nome)

        candidates = sorted(list(docs_dir.glob("*.doc")) + list(docs_dir.glob("*.docx")))
        for path in candidates:
            base_n = self._normalize_text_for_match(path.stem)
            if "autoriz" not in base_n:
                continue
            if cognome_n and cognome_n not in base_n:
                continue
            if nome_n and nome_n not in base_n:
                continue
            if numero and numero not in base_n:
                continue
            if year and year not in base_n:
                continue
            return path

        for path in candidates:
            base_n = self._normalize_text_for_match(path.stem)
            if "autoriz" in base_n and numero and numero in base_n:
                return path
        return None

    def apri_autorizzazione_record(self, record: dict):
        path = self._find_authorization_doc_for_record(record)
        if path is None:
            messagebox.showinfo(
                "Autorizzazione non trovata",
                f"Nessun file autorizzazione trovato in:\n{AUTH_DOCS_DIR}",
            )
            return
        try:
            os.startfile(str(path))
        except OSError as exc:
            messagebox.showerror("Apertura non riuscita", f"Impossibile aprire il file.\n\nDettagli:\n{exc}")

    def _generate_authorization_doc_with_word(self, template_path: Path, output_path: Path, payload: dict):
        ps_script = r"""
param(
    [Parameter(Mandatory=$true)][string]$TemplatePath,
    [Parameter(Mandatory=$true)][string]$OutputPath,
    [Parameter(Mandatory=$true)][string]$AuthNumber,
    [Parameter(Mandatory=$true)][string]$Year,
    [Parameter(Mandatory=$true)][string]$ReleaseDate,
    [Parameter(Mandatory=$true)][string]$ExpiryDate,
    [Parameter(Mandatory=$true)][string]$ProtocolNumber,
    [Parameter(Mandatory=$true)][string]$ProtocolDate,
    [Parameter(Mandatory=$true)][string]$Sigla,
    [Parameter(Mandatory=$true)][string]$Cognome,
    [Parameter(Mandatory=$true)][string]$Nome,
    [Parameter(Mandatory=$true)][string]$NatoWord,
    [Parameter(Mandatory=$true)][string]$BirthPlace,
    [Parameter(Mandatory=$true)][string]$BirthDate,
    [Parameter(Mandatory=$true)][string]$ResidenceCity,
    [Parameter(Mandatory=$true)][string]$ResidenceAddress,
    [Parameter(Mandatory=$true)][string]$IssueCity,
    [Parameter(Mandatory=$true)][string]$IssueDate
)
$ErrorActionPreference = "Stop"
Copy-Item -LiteralPath $TemplatePath -Destination $OutputPath -Force
$word = $null
$doc = $null
try {
    $word = New-Object -ComObject Word.Application
    $word.Visible = $false
    $doc = $word.Documents.Open($OutputPath)

    function Set-ParagraphLine([object]$Doc, [string]$Pattern, [string]$NewText) {
        foreach ($par in $Doc.Paragraphs) {
            $line = [string]$par.Range.Text
            $line = $line -replace "[`r`a]", ""
            $line = $line.Trim()
            if ($line -match $Pattern) {
                $par.Range.Text = $NewText + "`r"
                return $par
            }
        }
        return $null
    }

    $null = Set-ParagraphLine $doc "AUTORIZZAZIONE\s*N\." ("AUTORIZZAZIONE N. {0}/{1}" -f $AuthNumber, $Year)
    $null = Set-ParagraphLine $doc "Rilascio del .*Validit" ("Rilascio del {0} Validita sino {1}" -f $ReleaseDate, $ExpiryDate)
    $vistaPar = Set-ParagraphLine $doc "Vista la richiesta prot\." ("Vista la richiesta prot. n. {0} del {1}" -f $ProtocolNumber, $ProtocolDate)
    if ($vistaPar -ne $null) {
        $vistaPar.Range.Font.Bold = 0
        $start = $vistaPar.Range.Start
        $vistaRange = $doc.Range($start, $start + 5)
        $vistaRange.Font.Bold = 1
    }
    $intro = if ($Sigla -eq "Sig.ra") { "Alla" } else { "Al" }
    $sigLine = ("{0} {1} {2} {3} {4} a {5} il {6} residente a {7} in {8}" -f $intro, $Sigla, $Cognome, $Nome, $NatoWord, $BirthPlace, $BirthDate, $ResidenceCity, $ResidenceAddress)
    $null = Set-ParagraphLine $doc "^(Al|Alla)\s+(Sig\.|Sig\.ra|Sig.ra|Sig)\b" $sigLine
    $null = Set-ParagraphLine $doc "PEGOGNAGA\s+il" ("{0} il {1}" -f $IssueCity.ToUpper(), $IssueDate)

    $doc.Save()
}
finally {
    if ($doc -ne $null) { $doc.Close() }
    if ($word -ne $null) { $word.Quit() }
}
"""
        with tempfile.TemporaryDirectory(prefix="passapp_auth_") as tmp_dir:
            tmp_path = Path(tmp_dir)
            script_path = tmp_path / "create_auth.ps1"
            script_path.write_text(ps_script, encoding="utf-8")
            args = [
                "powershell",
                "-NoProfile",
                "-ExecutionPolicy",
                "Bypass",
                "-File",
                str(script_path),
                "-TemplatePath", str(template_path),
                "-OutputPath", str(output_path),
                "-AuthNumber", str(payload["auth_number"]),
                "-Year", str(payload["year"]),
                "-ReleaseDate", str(payload["release_date"]),
                "-ExpiryDate", str(payload["expiry_date"]),
                "-ProtocolNumber", str(payload["protocol_number"]),
                "-ProtocolDate", str(payload["protocol_date"]),
                "-Sigla", str(payload["sigla"]),
                "-Cognome", str(payload["cognome"]),
                "-Nome", str(payload["nome"]),
                "-NatoWord", str(payload["nato_word"]),
                "-BirthPlace", str(payload["birth_place"]),
                "-BirthDate", str(payload["birth_date"]),
                "-ResidenceCity", str(payload["residence_city"]),
                "-ResidenceAddress", str(payload["residence_address"]),
                "-IssueCity", str(payload["issue_city"]),
                "-IssueDate", str(payload["issue_date"]),
            ]
            result = subprocess.run(
                args,
                capture_output=True,
                text=True,
                encoding="utf-8",
                errors="replace",
                timeout=180,
            )
            if result.returncode != 0:
                details = (result.stderr or result.stdout or "Errore sconosciuto").strip()
                raise RuntimeError(details)

    def crea_autorizzazione_record(self, record: dict):
        template_path = self._resolve_authorization_template()
        if template_path is None:
            messagebox.showerror(
                "Template non trovato",
                f"Nessun template autorizzazione disponibile.\nControlla il percorso:\n{AUTH_TEMPLATE_PATH}",
            )
            return

        docs_dir = Path(AUTH_DOCS_DIR)
        if not docs_dir.exists():
            messagebox.showerror("Cartella non trovata", f"Percorso non disponibile:\n{AUTH_DOCS_DIR}")
            return

        cognome_default, nome_default = self._split_cognome_nome(record.get("nome", ""))
        protocol_date_default = parse_date(record.get("rilascio")) or datetime.date.today()
        expiry_default = parse_date(record.get("scadenza")) or protocol_date_default

        win = tk.Toplevel(self)
        win.title("Crea autorizzazione pass invalidi")
        win.configure(bg=BG)
        win.resizable(False, False)
        win.grab_set()

        body = tk.Frame(win, bg=SURFACE, highlightbackground=BORDER, highlightthickness=1)
        body.pack(fill="both", expand=True, padx=14, pady=14)

        vars_map = {
            "protocollo_numero": tk.StringVar(),
            "protocollo_data": tk.StringVar(value=protocol_date_default.strftime("%d/%m/%Y")),
            "sesso": tk.StringVar(value="F"),
            "cognome": tk.StringVar(value=cognome_default),
            "nome": tk.StringVar(value=nome_default),
            "luogo_nascita": tk.StringVar(),
            "data_nascita": tk.StringVar(value=format_date(record.get("nato_il", ""))),
            "comune_residenza": tk.StringVar(value="Pegognaga"),
            "indirizzo_residenza": tk.StringVar(value=str(record.get("indirizzo", "")).strip()),
            "data_scadenza": tk.StringVar(value=expiry_default.strftime("%d/%m/%Y")),
            "comune_emissione": tk.StringVar(value=AUTH_CITY_DEFAULT),
        }

        fields = [
            ("Numero protocollo", "protocollo_numero"),
            ("Data protocollo (GG/MM/AAAA)", "protocollo_data"),
            ("Cognome", "cognome"),
            ("Nome", "nome"),
            ("Luogo di nascita (es. Citta (PR))", "luogo_nascita"),
            ("Data di nascita (GG/MM/AAAA)", "data_nascita"),
            ("Comune di residenza", "comune_residenza"),
            ("Indirizzo di residenza", "indirizzo_residenza"),
            ("Scadenza autorizzazione (GG/MM/AAAA)", "data_scadenza"),
            ("Comune emissione", "comune_emissione"),
        ]

        def add_row(label_text, key):
            row = tk.Frame(body, bg=SURFACE)
            row.pack(fill="x", padx=12, pady=5)
            tk.Label(
                row,
                text=label_text,
                bg=SURFACE,
                fg=TEXT_MUTED,
                font=("Segoe UI", 9, "bold"),
                width=36,
                anchor="w",
            ).pack(side="left")
            tk.Entry(
                row,
                textvariable=vars_map[key],
                font=("Segoe UI", 10),
                bg="white",
                fg=TEXT,
                relief="solid",
                bd=1,
            ).pack(side="left", fill="x", expand=True, ipady=4)

        add_row(fields[0][0], fields[0][1])
        add_row(fields[1][0], fields[1][1])

        sex_row = tk.Frame(body, bg=SURFACE)
        sex_row.pack(fill="x", padx=12, pady=5)
        tk.Label(
            sex_row,
            text="Sesso",
            bg=SURFACE,
            fg=TEXT_MUTED,
            font=("Segoe UI", 9, "bold"),
            width=36,
            anchor="w",
        ).pack(side="left")
        ttk.Combobox(
            sex_row,
            textvariable=vars_map["sesso"],
            values=("F", "M"),
            state="readonly",
            width=10,
            font=("Segoe UI", 10),
        ).pack(side="left", fill="x", expand=True)

        for label_text, key in fields[2:]:
            add_row(label_text, key)

        def refresh_expiry(*_args):
            p_date = parse_date(vars_map["protocollo_data"].get().strip())
            b_date = parse_date(vars_map["data_nascita"].get().strip())
            if p_date is None or b_date is None:
                return
            expiry = self._compute_authorization_expiry(p_date, b_date)
            vars_map["data_scadenza"].set(expiry.strftime("%d/%m/%Y"))

        vars_map["protocollo_data"].trace_add("write", refresh_expiry)
        vars_map["data_nascita"].trace_add("write", refresh_expiry)

        footer = tk.Frame(body, bg=SURFACE)
        footer.pack(fill="x", padx=12, pady=(8, 12))

        def conferma():
            protocol_number = vars_map["protocollo_numero"].get().strip()
            protocol_date = parse_date(vars_map["protocollo_data"].get().strip())
            sesso = vars_map["sesso"].get().strip().upper()
            cognome = vars_map["cognome"].get().strip().upper()
            nome = vars_map["nome"].get().strip().title()
            luogo_nascita = vars_map["luogo_nascita"].get().strip()
            data_nascita = parse_date(vars_map["data_nascita"].get().strip())
            comune_res = vars_map["comune_residenza"].get().strip()
            indirizzo_res = vars_map["indirizzo_residenza"].get().strip()
            data_scadenza = parse_date(vars_map["data_scadenza"].get().strip())
            comune_emissione = vars_map["comune_emissione"].get().strip() or AUTH_CITY_DEFAULT

            if not protocol_number:
                messagebox.showwarning("Dati incompleti", "Inserisci il numero protocollo.", parent=win)
                return
            if protocol_date is None:
                messagebox.showwarning("Data non valida", "Data protocollo non valida.", parent=win)
                return
            if sesso not in {"F", "M"}:
                messagebox.showwarning("Valore non valido", "Seleziona il sesso (F/M).", parent=win)
                return
            if not cognome or not nome:
                messagebox.showwarning("Dati incompleti", "Inserisci cognome e nome.", parent=win)
                return
            if not luogo_nascita:
                messagebox.showwarning("Dati incompleti", "Inserisci il luogo di nascita.", parent=win)
                return
            if data_nascita is None:
                messagebox.showwarning("Data non valida", "Data di nascita non valida.", parent=win)
                return
            if not comune_res or not indirizzo_res:
                messagebox.showwarning("Dati incompleti", "Inserisci residenza completa.", parent=win)
                return
            if data_scadenza is None:
                messagebox.showwarning("Data non valida", "Data di scadenza non valida.", parent=win)
                return

            numero = self._record_numero_int(record)
            if numero is None:
                messagebox.showerror("Errore record", "Numero autorizzazione non valido.", parent=win)
                return

            year = protocol_date.year
            output_path = docs_dir / f"{cognome} {nome} Autoriz. disabili n. {numero} {year}{template_path.suffix}"
            if output_path.exists():
                if not messagebox.askyesno(
                    "File gia esistente",
                    f"Il file esiste gia:\n{output_path}\n\nVuoi sovrascriverlo?",
                    parent=win,
                ):
                    return

            payload = {
                "auth_number": numero,
                "year": year,
                "release_date": protocol_date.strftime("%d/%m/%Y"),
                "expiry_date": data_scadenza.strftime("%d/%m/%Y"),
                "protocol_number": protocol_number,
                "protocol_date": protocol_date.strftime("%d/%m/%Y"),
                "sigla": "Sig.ra" if sesso == "F" else "Sig.",
                "cognome": cognome,
                "nome": nome,
                "nato_word": "nata" if sesso == "F" else "nato",
                "birth_place": luogo_nascita,
                "birth_date": data_nascita.strftime("%d/%m/%Y"),
                "residence_city": comune_res,
                "residence_address": indirizzo_res,
                "issue_city": comune_emissione,
                "issue_date": datetime.date.today().strftime("%d/%m/%Y"),
            }

            try:
                self._generate_authorization_doc_with_word(template_path, output_path, payload)
            except Exception as exc:
                logger.exception("Errore creazione autorizzazione pass invalidi")
                messagebox.showerror("Creazione non riuscita", f"Impossibile creare il file .doc.\n\nDettagli:\n{exc}", parent=win)
                return

            win.destroy()
            messagebox.showinfo("Autorizzazione creata", f"File creato:\n{output_path}")
            try:
                os.startfile(str(output_path))
            except OSError:
                pass

        tk.Button(
            footer,
            text="Crea autorizzazione",
            bg=ACCENT,
            fg="white",
            font=("Segoe UI", 10, "bold"),
            relief="flat",
            cursor="hand2",
            activebackground=ACCENT_DARK,
            padx=14,
            pady=8,
            command=conferma,
        ).pack(side="left")
        tk.Button(
            footer,
            text="Annulla",
            bg=BG2,
            fg=TEXT_MUTED,
            font=("Segoe UI", 10, "bold"),
            relief="flat",
            cursor="hand2",
            activebackground=BORDER,
            highlightbackground=BORDER,
            highlightthickness=1,
            padx=14,
            pady=8,
            command=win.destroy,
        ).pack(side="left", padx=(8, 0))

    def mostra_dettaglio(self, event):
        record = self._selected_record()
        if record is None:
            return

        win = tk.Toplevel(self)
        win.title(f"Dettaglio - {record.get('nome', '-')}")
        win.geometry("640x560")
        win.configure(bg=BG)
        win.grab_set()

        content = tk.Frame(win, bg=BG)
        content.pack(fill="both", expand=True, padx=18, pady=18)

        fields = [
            ("Numero", str(record.get("numero", "-"))),
            ("Cognome e Nome", str(record.get("nome", "-"))),
            ("Nato il", format_date(record.get("nato_il"))),
            ("Indirizzo", str(record.get("indirizzo", "-"))),
            ("Rilascio", format_date(record.get("rilascio"))),
            ("Scadenza", format_date(record.get("scadenza"))),
            ("Note", str(record.get("note", "-"))),
            ("File", str(record.get("source", "-"))),
        ]
        for label, value in fields:
            row = tk.Frame(content, bg=SURFACE, highlightbackground=BORDER, highlightthickness=1)
            row.pack(fill="x", pady=3)
            tk.Label(row, text=label, bg=SURFACE, fg=TEXT_DIM, font=("Segoe UI", 9, "bold"), width=20, anchor="w", padx=10, pady=8).pack(side="left")
            tk.Label(row, text=value if value else "-", bg=SURFACE, fg=TEXT, font=("Segoe UI", 10), anchor="w", justify="left", wraplength=360, padx=10, pady=8).pack(side="left", fill="x", expand=True)

        footer = tk.Frame(win, bg=BG)
        footer.pack(fill="x", padx=18, pady=(0, 14))
        tk.Button(
            footer,
            text="Modifica dati",
            bg=ACCENT,
            fg="white",
            font=("Segoe UI", 10, "bold"),
            relief="flat",
            cursor="hand2",
            activebackground=ACCENT_DARK,
            padx=12,
            pady=8,
            command=lambda: (win.destroy(), self.modifica_selezionato()),
        ).pack(side="left")
        tk.Button(
            footer,
            text="Apri autorizzazione .doc",
            bg=SUCCESS,
            fg="white",
            font=("Segoe UI", 10, "bold"),
            relief="flat",
            cursor="hand2",
            activebackground="#197A52",
            padx=12,
            pady=8,
            command=lambda: self.apri_autorizzazione_record(record),
        ).pack(side="left", padx=(8, 0))
        tk.Button(
            footer,
            text="Crea autorizzazione",
            bg=WARNING,
            fg="white",
            font=("Segoe UI", 10, "bold"),
            relief="flat",
            cursor="hand2",
            activebackground="#9A5E0C",
            padx=12,
            pady=8,
            command=lambda: self.crea_autorizzazione_record(record),
        ).pack(side="left", padx=(8, 0))
        tk.Button(
            footer,
            text="Chiudi",
            bg=BG2,
            fg=TEXT_MUTED,
            font=("Segoe UI", 10, "bold"),
            relief="flat",
            cursor="hand2",
            activebackground=BORDER,
            highlightbackground=BORDER,
            highlightthickness=1,
            padx=12,
            pady=8,
            command=win.destroy,
        ).pack(side="right")

